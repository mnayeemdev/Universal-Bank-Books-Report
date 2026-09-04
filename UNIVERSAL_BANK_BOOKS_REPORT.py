"""
============================================================
UNIVERSAL BANK BOOKS REPORT V8.5.2
============================================================
Author   : Mohamed Nayeem
Copyright: © Mohamed Nayeem — All Rights Reserved
Instagram: @mohamednayeem7
============================================================

V8.5.2 — AUDIT MODES + CHARGES — HARDENED
----------------------------
Supports ALL major Indian bank statement formats:
  SBI, HDFC, ICICI, Axis, Kotak Mahindra, PNB, BOB,
  Canara, Union Bank, IDBI, Yes Bank, IndusInd, Federal,
  RBL, Bandhan, AU Small Finance, Ujjivan, Equitas,
  Jana, Suryoday, Utkarsh, and any bank that produces
  a standard tabular statement.

File types:  .xlsx  .xls  .csv  .pdf  .html  .htm  .txt

Outputs:
  BOOKS_OF_ACCOUNTS_V7_3_3.xlsx
  BANK_AUDIT_TECHNICAL_V7_3_3.xlsx

Sheets:
  COVER_REPORT
  ACCOUNT_PROFILE
  TRANSACTION_REGISTER
  REVIEW_REQUIRED
  DATE_WISE_SUMMARY
  MONTH_WISE_SUMMARY
  ACCOUNT_WISE_SUMMARY
  CUSTOMER_WISE_SUMMARY
  CUSTOMER_DATE_WISE
  TRANSACTION_TYPE_SUMMARY
  BANK_CHARGES_GST
  UTR_WISE_REGISTER
  UTR_DUPLICATE_CHECK
  BANK_RECONCILIATION
  PDF_PAGE_RECON
  DATA_QUALITY
  CONTROL_TOTALS
  SOURCE_MAPPING
  REPORT_NOTES

Core fields per transaction:
  Date, Customer Name, UTR / Reference, Narration,
  Debit, Credit, Balance, Direction, Transaction Amount,
  Customer Source, UTR Source, Possible Duplicate,
  Parser, Source File, Source Part, Source Page, Source Row

DATA INTEGRITY
  - No transaction is invented.
  - No missing name, UTR, debit, or credit is manufactured.
  - Customer/UTR extraction from narration is best-effort
    and explicitly marked.
  - Possible duplicates are flagged, never deleted.
  - PDF page totals are used as a control check.

INSTALL
  python -m pip install pandas openpyxl xlrd pdfplumber pyyaml lxml
Optional OCR:
  python -m pip install pytesseract pdf2image pillow

RUN
  python UNIVERSAL_BANK_BOOKS_REPORT_V8_5_2_AUDIT_MODES_CHARGES_HARDENED.py
"""

import os
import sys
import glob
import re
import yaml
import logging
import traceback
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# AUTHOR / BRANDING
# ============================================================
AUTHOR_NAME = "Mohamed Nayeem"
AUTHOR_COPYRIGHT = "© Mohamed Nayeem — All Rights Reserved"
AUTHOR_INSTAGRAM = "@mohamednayeem7"
VERSION = "V8_5_2_AUDIT_MODES_CHARGES_HARDENED"
DISPLAY_VERSION = "V8.5.2 AUDIT MODES + CHARGES — HARDENED"

VALIDATED_BANK_FORMATS = [
    "Axis Bank",
    "HDFC Bank",
    "ICICI Bank",
    "IDFC FIRST Bank",
    "South Indian Bank",
    "Standard Chartered Bank",
    "Yes Bank",
]

# V8.4 validation notes:
# - Tested against the supplied real XLSX bank statements.
# - Supports repeated page headers / shifted table columns.
# - ICICI SR283403308 multi-header statement no longer stops at first section.
# - Standard Chartered transactions after repeated headers are retained.
# - Genuine transactions are not removed merely because values repeat.

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FULL_OUTPUT_FILE = os.path.join(BASE_DIR, "UNIVERSAL_BANK_BOOKS_REPORT_V8_5_2_AUDIT_MODES_CHARGES_HARDENED.xlsx")
BOOKS_OUTPUT_FILE = FULL_OUTPUT_FILE
TECHNICAL_OUTPUT_FILE = FULL_OUTPUT_FILE
OUTPUT_FILE = FULL_OUTPUT_FILE
PATTERNS = (
    "*.xlsx", "*.XLSX",
    "*.xls",  "*.XLS",
    "*.csv",  "*.CSV",
    "*.pdf",  "*.PDF",
    "*.html", "*.HTML", "*.htm", "*.HTM",
    "*.txt", "*.TXT",
)
input_files = []
for pattern in PATTERNS:
    input_files.extend(glob.glob(os.path.join(BASE_DIR, pattern)))
input_files = sorted({
    p for p in input_files
    if os.path.abspath(p).lower() not in {
        os.path.abspath(FULL_OUTPUT_FILE).lower(),
        os.path.abspath(BOOKS_OUTPUT_FILE).lower(),
        os.path.abspath(TECHNICAL_OUTPUT_FILE).lower(),
    }
    and os.path.basename(p).lower() != "universal_bank_books_report_v8_simple.xlsx"
    and not os.path.basename(p).startswith("~$")
    and not re.match(
        r"(?i)^(?:UNIVERSAL_BANK_BOOKS_REPORT_V[^/\\]*|BOOKS_OF_ACCOUNTS_V[^/\\]*|BANK_AUDIT_TECHNICAL_V[^/\\]*)\.xlsx$",
        os.path.basename(p),
    )
    and not re.match(r"(?i)^audit_log_\d{8}_\d{6}\.txt$", os.path.basename(p))
})
if not input_files and "--self-test" not in sys.argv:
    raise FileNotFoundError(
        "No XLSX/XLS/CSV/PDF statement found in:\n" + BASE_DIR
    )

# ============================================================
# EXCEL STYLE
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
AUTHOR_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '₹#,##0.00;[Red](₹#,##0.00);-'
DATE_FMT = "dd-mm-yyyy"
COUNT_FMT = "#,##0"

# ============================================================
# BANK DETECTION SIGNATURES  (PDF text fingerprints)
# ============================================================
BANK_SIGNATURES = {
    "ICICI Bank": [
        "icici bank", "icici", "tran date", "value date",
        "particulars", "withdrawals", "deposits",
    ],
    "State Bank of India (SBI)": [
        "state bank of india", "sbi", "txn date",
        "value date", "description", "ref no",
        "debit", "credit", "balance",
    ],
    "HDFC Bank": [
        "hdfc bank", "hdfc", "narration",
        "chq./ref.no", "value dt", "withdrawal amt",
        "deposit amt", "closing balance",
    ],
    "Axis Bank": [
        "axis bank", "axis", "particulars",
        "cheque no", "debit", "credit", "balance",
    ],
    "Kotak Mahindra Bank": [
        "kotak", "transaction date", "value date",
        "transaction particulars", "cheque number",
        "debit", "credit", "balance",
    ],
    "Punjab National Bank (PNB)": [
        "punjab national bank", "pnb", "txn date",
        "value date", "particulars", "debit",
        "credit", "balance",
    ],
    "Bank of Baroda (BOB)": [
        "bank of baroda", "bob", "transaction date",
        "particulars", "cheque no", "debit",
        "credit", "balance",
    ],
    "Canara Bank": [
        "canara bank", "txn date", "value date",
        "particulars", "debit", "credit", "balance",
    ],
    "Union Bank of India": [
        "union bank", "transaction date", "value date",
        "particulars", "cheque no", "debit",
        "credit", "balance",
    ],
    "IDBI Bank": [
        "idbi bank", "idbi", "transaction date",
        "value date", "particulars", "debit",
        "credit", "balance",
    ],
    "Yes Bank": [
        "yes bank", "transaction date", "value date",
        "particulars", "ref no", "debit",
        "credit", "balance",
    ],
    "IndusInd Bank": [
        "indusind", "transaction date", "value date",
        "particulars", "cheque no", "debit",
        "credit", "balance",
    ],
    "Federal Bank": [
        "federal bank", "transaction date", "value date",
        "particulars", "ref no", "debit",
        "credit", "balance",
    ],
    "RBL Bank": [
        "rbl bank", "rbl", "transaction date",
        "particulars", "ref no", "debit",
        "credit", "balance",
    ],
    "Bandhan Bank": [
        "bandhan bank", "bandhan", "transaction date",
        "particulars", "debit", "credit", "balance",
    ],
    "AU Small Finance Bank": [
        "au small finance", "au bank",
        "transaction date", "particulars",
        "debit", "credit", "balance",
    ],
}

# V7.2 — extended India-wide bank fingerprints. Generic parsing still works
# for an unknown/co-operative bank when standard statement fields are present.
BANK_SIGNATURES.update({
    "IDFC FIRST Bank": ["idfc first", "idfcfirst", "transaction date", "balance"],
    "Indian Bank": ["indian bank", "transaction date", "particulars", "balance"],
    "Indian Overseas Bank": ["indian overseas bank", "iob", "transaction date", "balance"],
    "Bank of India": ["bank of india", "transaction date", "particulars", "balance"],
    "Central Bank of India": ["central bank of india", "transaction date", "balance"],
    "UCO Bank": ["uco bank", "transaction date", "balance"],
    "Bank of Maharashtra": ["bank of maharashtra", "transaction date", "balance"],
    "Punjab & Sind Bank": ["punjab & sind bank", "punjab and sind bank", "transaction date"],
    "Karur Vysya Bank": ["karur vysya bank", "kvb", "transaction date", "balance"],
    "City Union Bank": ["city union bank", "cub", "transaction date", "balance"],
    "Tamilnad Mercantile Bank": ["tamilnad mercantile bank", "tmb", "transaction date", "balance"],
    "South Indian Bank": ["south indian bank", "sib", "transaction date", "balance"],
    "Karnataka Bank": ["karnataka bank", "transaction date", "balance"],
    "DCB Bank": ["dcb bank", "transaction date", "balance"],
    "CSB Bank": ["csb bank", "catholic syrian bank", "transaction date"],
    "Dhanlaxmi Bank": ["dhanlaxmi bank", "transaction date", "balance"],
    "Jammu & Kashmir Bank": ["jammu & kashmir bank", "j&k bank", "jk bank"],
    "Nainital Bank": ["nainital bank", "transaction date", "balance"],
    "Jana Small Finance Bank": ["jana small finance bank", "jana", "transaction date"],
    "Suryoday Small Finance Bank": ["suryoday small finance bank", "suryoday", "transaction date"],
    "Utkarsh Small Finance Bank": ["utkarsh small finance bank", "utkarsh", "transaction date"],
    "Ujjivan Small Finance Bank": ["ujjivan small finance bank", "ujjivan", "transaction date"],
    "Equitas Small Finance Bank": ["equitas small finance bank", "equitas", "transaction date"],
    "ESAF Small Finance Bank": ["esaf small finance bank", "esaf", "transaction date"],
    "Capital Small Finance Bank": ["capital small finance bank", "capital sfb", "transaction date"],
    "Unity Small Finance Bank": ["unity small finance bank", "unity sfb", "transaction date"],
    "Shivalik Small Finance Bank": ["shivalik small finance bank", "shivalik", "transaction date"],
    "Airtel Payments Bank": ["airtel payments bank", "airtel bank", "transaction date"],
    "India Post Payments Bank": ["india post payments bank", "ippb", "transaction date"],
    "Fino Payments Bank": ["fino payments bank", "fino bank", "transaction date"],
    "Jio Payments Bank": ["jio payments bank", "jio bank", "transaction date"],
    "NSDL Payments Bank": ["nsdl payments bank", "nsdl bank", "transaction date"],
    # Foreign banks operating in India / India banking operations
    "Standard Chartered Bank": ["standard chartered", "scb", "transaction date", "account statement"],
    "HSBC": ["hsbc", "hongkong and shanghai banking corporation", "account statement"],
    "Citibank N.A.": ["citibank", "citi bank", "citibank n.a", "account statement"],
    "Deutsche Bank": ["deutsche bank", "transaction details", "account statement"],
    "DBS Bank India": ["dbs bank india", "dbs bank", "digibank", "account statement"],
    "Bank of America": ["bank of america", "bofa", "account statement"],
    "Barclays Bank": ["barclays bank", "barclays", "account statement"],
    "BNP Paribas": ["bnp paribas", "account statement"],
    "JPMorgan Chase Bank": ["jpmorgan chase", "jp morgan", "jpmorgan", "account statement"],
    "MUFG Bank": ["mufg bank", "bank of tokyo-mitsubishi ufj", "account statement"],
    "Mizuho Bank": ["mizuho bank", "account statement"],
    "Sumitomo Mitsui Banking Corporation": ["sumitomo mitsui banking corporation", "smbc", "account statement"],
    "Bank of Bahrain & Kuwait": ["bank of bahrain & kuwait", "bank of bahrain and kuwait", "bbk", "account statement"],
    "Doha Bank": ["doha bank", "account statement"],
    "Emirates NBD": ["emirates nbd", "enbd", "account statement"],
    "First Abu Dhabi Bank": ["first abu dhabi bank", "fab bank", "account statement"],
    "Mashreq Bank": ["mashreq bank", "mashreq", "account statement"],
    "Qatar National Bank": ["qatar national bank", "qnb", "account statement"],
    "Société Générale": ["société générale", "societe generale", "account statement"],
    "Bank of China": ["bank of china", "account statement"],
    "Industrial and Commercial Bank of China": ["industrial and commercial bank of china", "icbc", "account statement"],
    "KEB Hana Bank": ["keb hana bank", "hana bank", "account statement"],
    "Shinhan Bank": ["shinhan bank", "account statement"],
    "Woori Bank": ["woori bank", "account statement"],
})

# Official V7.2 catalogue used for reporting/category tagging.
BANK_CATALOG = {
    # Public Sector Banks
    "State Bank of India (SBI)": "PUBLIC SECTOR BANK",
    "Bank of Baroda (BOB)": "PUBLIC SECTOR BANK",
    "Bank of India": "PUBLIC SECTOR BANK",
    "Bank of Maharashtra": "PUBLIC SECTOR BANK",
    "Canara Bank": "PUBLIC SECTOR BANK",
    "Central Bank of India": "PUBLIC SECTOR BANK",
    "Indian Bank": "PUBLIC SECTOR BANK",
    "Indian Overseas Bank": "PUBLIC SECTOR BANK",
    "Punjab National Bank (PNB)": "PUBLIC SECTOR BANK",
    "Punjab & Sind Bank": "PUBLIC SECTOR BANK",
    "UCO Bank": "PUBLIC SECTOR BANK",
    "Union Bank of India": "PUBLIC SECTOR BANK",
    # Private Sector Banks
    "Axis Bank": "PRIVATE SECTOR BANK", "Bandhan Bank": "PRIVATE SECTOR BANK",
    "CSB Bank": "PRIVATE SECTOR BANK", "City Union Bank": "PRIVATE SECTOR BANK",
    "DCB Bank": "PRIVATE SECTOR BANK", "Dhanlaxmi Bank": "PRIVATE SECTOR BANK",
    "Federal Bank": "PRIVATE SECTOR BANK", "HDFC Bank": "PRIVATE SECTOR BANK",
    "ICICI Bank": "PRIVATE SECTOR BANK", "IDBI Bank": "PRIVATE SECTOR BANK",
    "IDFC FIRST Bank": "PRIVATE SECTOR BANK", "IndusInd Bank": "PRIVATE SECTOR BANK",
    "Jammu & Kashmir Bank": "PRIVATE SECTOR BANK", "Karnataka Bank": "PRIVATE SECTOR BANK",
    "Karur Vysya Bank": "PRIVATE SECTOR BANK", "Kotak Mahindra Bank": "PRIVATE SECTOR BANK",
    "Nainital Bank": "PRIVATE SECTOR BANK", "RBL Bank": "PRIVATE SECTOR BANK",
    "South Indian Bank": "PRIVATE SECTOR BANK", "Tamilnad Mercantile Bank": "PRIVATE SECTOR BANK",
    "Yes Bank": "PRIVATE SECTOR BANK",
    # Small Finance Banks
    "AU Small Finance Bank": "SMALL FINANCE BANK", "Capital Small Finance Bank": "SMALL FINANCE BANK",
    "Equitas Small Finance Bank": "SMALL FINANCE BANK", "ESAF Small Finance Bank": "SMALL FINANCE BANK",
    "Jana Small Finance Bank": "SMALL FINANCE BANK", "Shivalik Small Finance Bank": "SMALL FINANCE BANK",
    "Suryoday Small Finance Bank": "SMALL FINANCE BANK", "Ujjivan Small Finance Bank": "SMALL FINANCE BANK",
    "Unity Small Finance Bank": "SMALL FINANCE BANK", "Utkarsh Small Finance Bank": "SMALL FINANCE BANK",
    # Payments Banks
    "Airtel Payments Bank": "PAYMENTS BANK", "India Post Payments Bank": "PAYMENTS BANK",
    "Fino Payments Bank": "PAYMENTS BANK", "Jio Payments Bank": "PAYMENTS BANK",
    "NSDL Payments Bank": "PAYMENTS BANK",
    # Foreign Banks
    "Standard Chartered Bank": "FOREIGN BANK", "HSBC": "FOREIGN BANK", "Citibank N.A.": "FOREIGN BANK",
    "Deutsche Bank": "FOREIGN BANK", "DBS Bank India": "FOREIGN BANK", "Bank of America": "FOREIGN BANK",
    "Barclays Bank": "FOREIGN BANK", "BNP Paribas": "FOREIGN BANK", "JPMorgan Chase Bank": "FOREIGN BANK",
    "MUFG Bank": "FOREIGN BANK", "Mizuho Bank": "FOREIGN BANK",
    "Sumitomo Mitsui Banking Corporation": "FOREIGN BANK", "Bank of Bahrain & Kuwait": "FOREIGN BANK",
    "Doha Bank": "FOREIGN BANK", "Emirates NBD": "FOREIGN BANK", "First Abu Dhabi Bank": "FOREIGN BANK",
    "Mashreq Bank": "FOREIGN BANK", "Qatar National Bank": "FOREIGN BANK", "Société Générale": "FOREIGN BANK",
    "Bank of China": "FOREIGN BANK", "Industrial and Commercial Bank of China": "FOREIGN BANK",
    "KEB Hana Bank": "FOREIGN BANK", "Shinhan Bank": "FOREIGN BANK", "Woori Bank": "FOREIGN BANK",
}

def canonical_bank_name(bank_name):
    """Normalize known spelling/display variants to BANK_CATALOG keys."""
    name = " ".join(str(bank_name or "").strip().split())
    aliases = {
        "Bank of Baroda": "Bank of Baroda (BOB)",
        "BOB": "Bank of Baroda (BOB)",
        "YES Bank": "Yes Bank",
        "YES BANK": "Yes Bank",
        "State Bank of India": "State Bank of India (SBI)",
        "SBI": "State Bank of India (SBI)",
        "Punjab National Bank": "Punjab National Bank (PNB)",
        "PNB": "Punjab National Bank (PNB)",
    }
    return aliases.get(name, name)


def bank_category(bank_name):
    name = canonical_bank_name(clean_text(bank_name))
    if name in BANK_CATALOG:
        return BANK_CATALOG[name]
    if name == "Unknown Bank":
        return "UNKNOWN / OTHER BANK"
    return "OTHER / CO-OPERATIVE / LEGACY BANK"


def detect_bank_from_text(text):
    """Best-effort bank detection using identity terms only.

    V7.2 deliberately ignores generic headers such as 'transaction date',
    'particulars', 'debit', 'credit' and 'balance' for bank identity. This
    prevents a generic statement from being falsely labelled as a specific bank.
    """
    t = clean_text(text).lower()
    if not t:
        return "Unknown Bank"

    # V7.3 FIX: identify the statement-owning bank from its IFSC before
    # scanning transaction narration.  Counterparty bank names inside the
    # first transaction rows must never override the statement bank.
    ifsc_owner = {
        "sbin": "State Bank of India (SBI)", "barb": "Bank of Baroda (BOB)",
        "bkid": "Bank of India", "mahb": "Bank of Maharashtra",
        "cnrb": "Canara Bank", "cbin": "Central Bank of India",
        "idib": "Indian Bank", "ioba": "Indian Overseas Bank",
        "punb": "Punjab National Bank (PNB)", "psib": "Punjab & Sind Bank",
        "ucba": "UCO Bank", "ubin": "Union Bank of India",
        "utib": "Axis Bank", "bdbl": "Bandhan Bank", "cbin": "Central Bank of India",
        "ciub": "City Union Bank", "dcbl": "DCB Bank", "dlxb": "Dhanlaxmi Bank",
        "fdrl": "Federal Bank", "hdfc": "HDFC Bank", "icic": "ICICI Bank",
        "ibkl": "IDBI Bank", "idfb": "IDFC FIRST Bank", "indb": "IndusInd Bank",
        "jaka": "Jammu & Kashmir Bank", "karb": "Karnataka Bank",
        "kvbl": "Karur Vysya Bank", "kkbk": "Kotak Mahindra Bank",
        "ratn": "RBL Bank", "sibl": "South Indian Bank", "tmbL".lower(): "Tamilnad Mercantile Bank",
        "yesb": "Yes Bank", "aubl": "AU Small Finance Bank", "esmf": "ESAF Small Finance Bank",
        "jsfb": "Jana Small Finance Bank", "sury": "Suryoday Small Finance Bank",
        "ujvn": "Ujjivan Small Finance Bank", "utks": "Utkarsh Small Finance Bank",
        "scbl": "Standard Chartered Bank", "hsbc": "HSBC", "dbss": "DBS Bank India",
    }
    header_zone = t[:6000]
    for match in re.finditer(r"\b([a-z]{4})0[a-z0-9]{6}\b", header_zone, flags=re.I):
        prefix = match.group(1).lower()
        if prefix in ifsc_owner:
            return ifsc_owner[prefix]

    generic_terms = {
        "transaction date", "tran date", "txn date", "value date",
        "particulars", "narration", "description", "transaction details",
        "debit", "credit", "balance", "withdrawals", "deposits",
        "withdrawal amt", "deposit amt", "closing balance", "ref no",
        "cheque no", "cheque number", "account statement",
    }
    candidates = []
    for bank, keywords in BANK_SIGNATURES.items():
        best = 0
        best_term = ""
        for kw in keywords:
            term = clean_text(kw).lower()
            if not term or term in generic_terms:
                continue
            # Very short abbreviations can create false matches inside words.
            if len(term) <= 4 and term.isalnum():
                matched = re.search(r"\b" + re.escape(term) + r"\b", t) is not None
            else:
                matched = term in t
            if matched:
                # Prefer longer/more distinctive identity phrases.
                score = len(term.replace(" ", ""))
                if "bank" in term:
                    score += 10
                if score > best:
                    best = score
                    best_term = term
        if best:
            candidates.append((best, len(best_term), bank))
    if not candidates:
        return "Unknown Bank"
    candidates.sort(reverse=True)
    return candidates[0][2]


# ============================================================
# GENERIC BANK HEADER ALIASES  (expanded for all banks)
# ============================================================
ALIASES = {
    "date": [
        "date", "tran date", "transaction date", "txn date",
        "trn date", "posting date", "value date",
        "date of transaction", "txn. date", "trans date",
        "transaction dt", "tr date", "book date",
        "effective date", "entry date",
    ],
    "narration": [
        "particulars", "narration", "description", "details",
        "trn particulars", "trn. particulars",
        "transaction particulars", "transaction description",
        "transaction details", "remarks", "transaction narrative",
        "txn narration", "txn description", "details of transaction",
        "particular", "desc", "narrative",
    ],
    "debit": [
        "debit", "debit amount", "debit amt",
        "withdrawal", "withdrawals", "withdrawal amount",
        "withdrawal amt", "dr", "dr amount", "paid out",
        "debits", "dr.", "debit(rs)", "debit (rs)",
        "debit (inr)", "debit(inr)", "withdrawal(rs)",
        "withdrawals(rs)", "outflow", "money out",
    ],
    "credit": [
        "credit", "credit amount", "credit amt",
        "deposit", "deposits", "deposit amount",
        "deposit amt", "cr", "cr amount", "paid in",
        "credits", "cr.", "credit(rs)", "credit (rs)",
        "credit (inr)", "credit(inr)", "deposit(rs)",
        "deposits(rs)", "inflow", "money in",
    ],
    "balance": [
        "balance", "balance inr", "balance (inr)",
        "closing balance", "running balance",
        "available balance", "ledger balance",
        "bal", "bal (inr)", "bal(inr)",
        "closing bal", "balance(rs)", "balance (rs)",
        "end balance", "closing",
    ],
    "customer": [
        "customer", "customer name", "counterparty",
        "party name", "payer", "payer name",
        "payee", "payee name", "beneficiary",
        "beneficiary name", "remitter name",
        "sender name", "receiver name", "party",
        "name", "counterparty name",
    ],
    "utr": [
        "utr", "utr no", "utr number", "bank utr",
        "reference", "reference no", "reference number",
        "reference id", "ref id", "bank ref",
        "bank ref no", "bank reference", "rrn",
        "txn id", "transaction id", "bank transaction id",
        "ref no", "ref no.", "chq./ref.no",
        "chq/ref no", "cheque no", "cheque no.",
        "cheque number", "chq no", "chq. no",
        "ref number", "transaction ref", "txn ref",
    ],
    "amount": [
        "amount", "transaction amount", "txn amount",
        "amt", "transaction amt", "txn amt",
    ],
    "type": [
        "type", "transaction type", "txn type",
        "dr cr", "cr dr", "debit credit",
        "credit debit", "d/c", "c/d", "dr/cr",
    ],
}


# V7.2 — extra header variants seen across Indian retail/current-account exports.
_EXTRA_ALIASES = {
    "date": ["txn dt", "transaction dt.", "posting dt", "date & time", "transaction date & time"],
    "narration": ["transaction remarks", "transaction narration", "description/narration", "particulars / narration", "txn particulars / narration"],
    "debit": ["withdrawal amt.", "withdrawal amount(inr)", "debit amount(inr)", "debit amt(inr)", "dr amt", "dr amt."],
    "credit": ["deposit amt.", "deposit amount(inr)", "credit amount(inr)", "credit amt(inr)", "cr amt", "cr amt."],
    "balance": ["closing balance(inr)", "available bal", "ledger bal", "running bal", "balance amount", "balance amt"],
    "customer": ["remitter", "beneficiary / remitter", "counter party", "counter-party", "transaction party"],
    "utr": ["utr/ref no", "utr / ref no", "reference/cheque no", "chq/ref", "cheque/reference", "transaction reference no", "rrn/utr"],
    "amount": ["txn amt.", "transaction value", "value amount", "debit/credit amount", "cr/dr amount"],
    "type": ["dr/cr indicator", "debit/credit indicator", "txn nature", "nature", "transaction nature", "cr/dr"],
}
for _field, _values in _EXTRA_ALIASES.items():
    _seen = {re.sub(r"[^a-z0-9]+", " ", str(v).lower()).strip() for v in ALIASES.setdefault(_field, [])}
    for _value in _values:
        _key = re.sub(r"[^a-z0-9]+", " ", str(_value).lower()).strip()
        if _key not in _seen:
            ALIASES[_field].append(_value)
            _seen.add(_key)

# ============================================================
# V7 CONFIGURATION — VALIDATED AND MERGED WITH BUILT-IN ALIASES
# ============================================================
# Important: config aliases EXTEND the proven V6/V7 built-in aliases.
# They never replace them, so a short custom YAML cannot reduce bank support.
CONFIG_CANDIDATES = [
    os.path.join(BASE_DIR, "config.yaml"),
    os.path.join(BASE_DIR, "UNIVERSAL_BANK_BOOKS_CONFIG_V7.yaml"),
]
CONFIG_FILE = next((p for p in CONFIG_CANDIDATES if os.path.exists(p)), CONFIG_CANDIDATES[0])

DEFAULT_CONFIG = {
    "column_aliases": {},
    "rounding_precision": 2,
    "max_rows_per_sheet": 1000000,
    "preserve_raw_data": True,
    "review_confidence_ok": 85,
    "review_confidence_min": 65,
    # PG/GST validation is deliberately disabled for ordinary bank statements.
    # A bank transaction alone does not prove that a PG charge applies.
    "pg_validation_enabled": False,
    "pg_charge_rate": 0.008,
    "gst_rate": 0.18,
    "success_statuses": ["SUCCESS", "COMPLETED", "CAPTURED", "SETTLED"],
    "failed_statuses": ["FAILED", "DECLINED", "REJECTED", "ERROR"],
    "pending_statuses": ["PENDING", "INPROCESS", "AUTHORIZED"],
    "diagnostic_mode": True,
    "ocr_enabled": True,
    "allow_leading_serial_before_date": True,
    "unknown_bank_generic_parser": True,
}


def validate_config(cfg):
    if cfg is None:
        cfg = {}
    if not isinstance(cfg, dict):
        raise ValueError("config.yaml root must be a mapping/object.")

    merged = dict(DEFAULT_CONFIG)
    merged.update(cfg)

    aliases_cfg = merged.get("column_aliases", {})
    if aliases_cfg is None:
        aliases_cfg = {}
    if not isinstance(aliases_cfg, dict):
        raise ValueError("column_aliases must be a mapping of field -> list of aliases.")
    for field, values in aliases_cfg.items():
        if not isinstance(field, str):
            raise ValueError("Every column_aliases key must be text.")
        if not isinstance(values, list) or not all(isinstance(v, str) for v in values):
            raise ValueError(f"column_aliases.{field} must be a list of strings.")

    try:
        merged["rounding_precision"] = int(merged.get("rounding_precision", 2))
    except Exception:
        raise ValueError("rounding_precision must be an integer.")
    if not 0 <= merged["rounding_precision"] <= 8:
        raise ValueError("rounding_precision must be between 0 and 8.")

    try:
        merged["max_rows_per_sheet"] = int(merged.get("max_rows_per_sheet", 1000000))
    except Exception:
        raise ValueError("max_rows_per_sheet must be an integer.")
    # Excel supports 1,048,576 rows including header.
    if not 1000 <= merged["max_rows_per_sheet"] <= 1048575:
        raise ValueError("max_rows_per_sheet must be between 1,000 and 1,048,575 data rows.")

    for key in ("review_confidence_ok", "review_confidence_min"):
        try:
            merged[key] = int(merged.get(key, DEFAULT_CONFIG[key]))
        except Exception:
            raise ValueError(f"{key} must be an integer.")
        if not 0 <= merged[key] <= 100:
            raise ValueError(f"{key} must be between 0 and 100.")
    if merged["review_confidence_min"] > merged["review_confidence_ok"]:
        raise ValueError("review_confidence_min cannot be greater than review_confidence_ok.")

    for key in ("pg_charge_rate", "gst_rate"):
        try:
            merged[key] = float(merged.get(key, DEFAULT_CONFIG[key]))
        except Exception:
            raise ValueError(f"{key} must be numeric.")
        if merged[key] < 0 or merged[key] > 1:
            raise ValueError(f"{key} must be between 0 and 1 (for example 0.008 = 0.8%).")

    merged["preserve_raw_data"] = bool(merged.get("preserve_raw_data", True))
    merged["pg_validation_enabled"] = bool(merged.get("pg_validation_enabled", False))
    merged["diagnostic_mode"] = bool(merged.get("diagnostic_mode", True))
    merged["ocr_enabled"] = bool(merged.get("ocr_enabled", True))
    merged["allow_leading_serial_before_date"] = bool(merged.get("allow_leading_serial_before_date", True))
    merged["unknown_bank_generic_parser"] = bool(merged.get("unknown_bank_generic_parser", True))

    for key in ("success_statuses", "failed_statuses", "pending_statuses"):
        values = merged.get(key, DEFAULT_CONFIG[key])
        if not isinstance(values, list):
            raise ValueError(f"{key} must be a list.")
        merged[key] = [str(v).strip().upper() for v in values if str(v).strip()]

    merged["column_aliases"] = aliases_cfg
    return merged


if not os.path.exists(CONFIG_FILE):
    # Create a safe default and continue. No forced exit is needed.
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        yaml.safe_dump(DEFAULT_CONFIG, f, sort_keys=False, allow_unicode=True)

try:
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        CONFIG = validate_config(yaml.safe_load(f))
except Exception as exc:
    raise ValueError(f"Invalid YAML configuration: {CONFIG_FILE}\\n{exc}") from exc

# Merge custom aliases into built-in aliases without deleting built-in support.
for field, custom_values in CONFIG.get("column_aliases", {}).items():
    existing = ALIASES.setdefault(field, [])
    seen = {str(v).strip().lower() for v in existing}
    for value in custom_values:
        key = str(value).strip().lower()
        if key and key not in seen:
            existing.append(value)
            seen.add(key)

ROUND_DECIMALS = CONFIG["rounding_precision"]
MAX_ROWS_PER_SHEET = CONFIG["max_rows_per_sheet"]
PRESERVE_RAW_DATA = True  # V8.2: user-required REAL RAW preservation; config cannot disable it
REVIEW_OK_THRESHOLD = CONFIG["review_confidence_ok"]
REVIEW_MIN_THRESHOLD = CONFIG["review_confidence_min"]
PG_VALIDATION_ENABLED = CONFIG["pg_validation_enabled"]
PG_CHARGE_RATE = CONFIG["pg_charge_rate"]
GST_RATE = CONFIG["gst_rate"]
DIAGNOSTIC_MODE = CONFIG["diagnostic_mode"]
OCR_ENABLED = CONFIG["ocr_enabled"]
ALLOW_LEADING_SERIAL_BEFORE_DATE = CONFIG["allow_leading_serial_before_date"]
UNKNOWN_BANK_GENERIC_PARSER = CONFIG["unknown_bank_generic_parser"]

# ============================================================
# V7 LOGGING / AUDIT TRAIL
# ============================================================
RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = os.path.join(BASE_DIR, f"audit_log_{RUN_ID}.txt")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("universal_bank_books_v7")
logger.info("Run ID: %s", RUN_ID)
logger.info("Config: %s", CONFIG_FILE)
logger.info("Output: %s", OUTPUT_FILE)
logger.info("Validated XLSX formats: %s", ", ".join(VALIDATED_BANK_FORMATS))
logger.info("Multi-header / shifted-column parsing: ENABLED")

exceptions_list = []
raw_data_sheets = {}
raw_data_index_records = []
row_control_seed = []
parser_recon_records = []
format_diagnostic_records = []


def log_exception(category, description, source_file="", source_part="", source_row=""):
    record = {
        "Exception Type": str(category),
        "Description": str(description),
        "Source File": str(source_file),
        "Source Part": str(source_part),
        "Source Row": source_row,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    exceptions_list.append(record)
    logger.warning("%s: %s | %s | %s", category, description, source_file, source_part)


def sanitize_sheet_name(name, used=None):
    value = re.sub(r"[\\/*?:\\[\\]]", "_", str(name or "Sheet"))
    value = value.strip(" '") or "Sheet"
    value = value[:31]
    if used is None:
        return value
    candidate = value
    n = 2
    while candidate.lower() in used:
        suffix = f"_{n}"
        candidate = value[:31-len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def store_raw_data(df, source_file, source_part):
    if not PRESERVE_RAW_DATA or df is None or df.empty:
        return
    chunk_size = min(MAX_ROWS_PER_SHEET, 1048575)
    total = len(df)
    for start in range(0, total, chunk_size):
        chunk = df.iloc[start:start + chunk_size].copy()
        sheet_name = f"RAW_{len(raw_data_sheets) + 1:03d}"
        raw_data_sheets[sheet_name] = chunk
        raw_data_index_records.append({
            "Raw Sheet": sheet_name,
            "Source File": source_file,
            "Source Part": source_part,
            "Start Record": start + 1,
            "End Record": min(start + chunk_size, total),
            "Rows": len(chunk),
        })

# ============================================================
# UNIVERSAL DATE PATTERNS  (all Indian bank date formats)
# ============================================================
UNIVERSAL_DATE_PATTERNS = [
    # DD-MM-YYYY  (ICICI, SBI, PNB, Canara)
    (re.compile(r"\b(\d{2})-(\d{2})-(\d{4})\b"), "%d-%m-%Y"),
    # DD/MM/YYYY  (HDFC, Axis, Kotak, BOB, Union, IDBI)
    (re.compile(r"\b(\d{2})/(\d{2})/(\d{4})\b"), "%d/%m/%Y"),
    # DD.MM.YYYY  (some banks)
    (re.compile(r"\b(\d{2})\.(\d{2})\.(\d{4})\b"), "%d.%m.%Y"),
    # DD-MMM-YYYY  (SBI old, PNB old)
    (re.compile(
        r"\b(\d{2})-(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{4})\b",
        re.I,
    ), "%d-%b-%Y"),
    # DD MMM YYYY
    (re.compile(
        r"\b(\d{2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})\b",
        re.I,
    ), "%d %b %Y"),
    # DD/MMM/YYYY
    (re.compile(
        r"\b(\d{2})/(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)/(\d{4})\b",
        re.I,
    ), "%d/%b/%Y"),
    # YYYY-MM-DD  (some digital banks / ISO)
    (re.compile(r"\b(\d{4})-(\d{2})-(\d{2})\b"), "%Y-%m-%d"),
    # DD-MM-YY  (older formats)
    (re.compile(r"\b(\d{2})-(\d{2})-(\d{2})\b"), "%d-%m-%y"),
    # DD/MM/YY
    (re.compile(r"\b(\d{2})/(\d{2})/(\d{2})\b"), "%d/%m/%y"),
    # DD.MM.YY
    (re.compile(r"\b(\d{2})\.(\d{2})\.(\d{2})\b"), "%d.%m.%y"),
    # DD-MMM-YY / DD MMM YY / DD/MMM/YY
    (re.compile(r"\b(\d{1,2})-(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{2})\b", re.I), "%d-%b-%y"),
    (re.compile(r"\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{2})\b", re.I), "%d %b %y"),
    (re.compile(r"\b(\d{1,2})/(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)/(\d{2})\b", re.I), "%d/%b/%y"),
    # YYYY/MM/DD
    (re.compile(r"\b(\d{4})/(\d{2})/(\d{2})\b"), "%Y/%m/%d"),
]

# Date-start regex for line detection (matches beginning of line)
DATE_LINE_PATTERNS = [
    re.compile(r"^(\d{2}-\d{2}-\d{4})\b"),
    re.compile(r"^(\d{2}/\d{2}/\d{4})\b"),
    re.compile(r"^(\d{2}\.\d{2}\.\d{4})\b"),
    re.compile(
        r"^(\d{2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{4})\b",
        re.I,
    ),
    re.compile(
        r"^(\d{2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})\b",
        re.I,
    ),
    re.compile(r"^(\d{4}-\d{2}-\d{2})\b"),
    re.compile(r"^(\d{2}-\d{2}-\d{2})\b"),
    re.compile(r"^(\d{2}/\d{2}/\d{2})\b"),
]


def parse_date_flexible(text):
    """Try all known date formats and return a datetime or NaT."""
    s = clean_text(text).strip()
    if not s:
        return pd.NaT
    for pattern, fmt in UNIVERSAL_DATE_PATTERNS:
        m = pattern.search(s)
        if m:
            try:
                return pd.to_datetime(m.group(0), format=fmt)
            except Exception:
                continue
    # Last resort: let pandas guess
    try:
        return pd.to_datetime(s, dayfirst=True, errors="coerce")
    except Exception:
        return pd.NaT


def match_date_at_line_start(line):
    """Return transaction date at line start; V7.2 tolerates a leading serial number."""
    text = str(line or "").strip()
    for pat in DATE_LINE_PATTERNS:
        m = pat.match(text)
        if m:
            return m.group(1)
    if ALLOW_LEADING_SERIAL_BEFORE_DATE:
        candidate = re.sub(r"^\s*\d{1,7}\s*(?:[|.:;-]\s*)?", "", text, count=1)
        for pat in DATE_LINE_PATTERNS:
            m = pat.match(candidate)
            if m:
                return m.group(1)
    return None


# ============================================================
# BASIC HELPERS
# ============================================================
def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    return " ".join(
        str(value)
        .replace("\ufeff", " ")
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .split()
    )


def norm(value):
    return re.sub(
        r"[^a-z0-9]+",
        " ",
        clean_text(value).lower(),
    ).strip()


def parse_number(value):
    if value is None:
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            return round(float(value), ROUND_DECIMALS)
        except Exception:
            return np.nan
    s = clean_text(value)
    if not s:
        return np.nan
    negative = s.startswith("(") and s.endswith(")")
    trailing_minus = s.endswith("-") and not s.startswith("-")
    if negative:
        s = s[1:-1]
    elif trailing_minus:
        s = s[:-1].strip()
        negative = True
    s = (
        s.replace(",", "")
        .replace("₹", "")
        .replace("INR", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .strip()
    )
    s = re.sub(r"\s*(DR|CR)\s*$", "", s, flags=re.I).strip()
    try:
        x = float(s)
        return round(-x if negative else x, ROUND_DECIMALS)
    except Exception:
        return np.nan


MONEY_TOKEN_RE = re.compile(
    r"(?<![\w])"
    r"(?:\d{1,3}(?:,\d{2,3})+|\d+)"
    r"(?:\.\d{2})?"
    r"(?![\w])"
)


def alias_match(value, field):
    v = norm(value)
    if not v:
        return False
    for alias in ALIASES[field]:
        a = norm(alias)
        if v == a:
            return True
        # Avoid false matches such as generic "Amount" being treated as
        # both "Debit Amount" and "Credit Amount". Prefer alias->header
        # containment; allow reverse containment only for descriptive headers.
        if len(a) >= 4 and a in v:
            return True
        if len(v) >= 8 and len(v.split()) >= 2 and v in a:
            return True
    return False


def find_column(columns, field):
    exact = {norm(c): c for c in columns}
    for alias in ALIASES[field]:
        a = norm(alias)
        if a in exact:
            return exact[a]
    for c in columns:
        if alias_match(c, field):
            return c
    return None


def header_hits(values):
    hits = set()
    for field in ALIASES:
        if any(alias_match(v, field) for v in values):
            hits.add(field)
    return hits


def is_generic_bank_header(values):
    hits = header_hits(values)
    if "date" in hits and ("debit" in hits or "credit" in hits):
        return True
    if "date" in hits and "amount" in hits and "type" in hits:
        return True
    # V7.2: many co-operative/digital bank exports keep DR/CR inside the
    # Amount cells and therefore have no separate Type column.
    if "date" in hits and "amount" in hits and ("balance" in hits or "narration" in hits):
        return True
    return False


def locate_generic_header(raw, max_rows=150):
    limit = min(len(raw), max_rows)
    best_row = None
    best_score = -1
    for i in range(limit):
        vals = raw.iloc[i].tolist()
        hits = header_hits(vals)
        if not is_generic_bank_header(vals):
            continue
        score = (
            5 * ("date" in hits)
            + 3 * ("debit" in hits)
            + 3 * ("credit" in hits)
            + 2 * ("narration" in hits)
            + 2 * ("amount" in hits)
            + 2 * ("type" in hits)
            + 1 * ("balance" in hits)
            + 1 * ("customer" in hits)
            + 1 * ("utr" in hits)
        )
        if score > best_score:
            best_score = score
            best_row = i
    return best_row



def locate_all_generic_headers(raw, max_rows=None):
    """
    V8.3 FIX:
    Find every valid bank-table header in a sheet, not only the first one.

    Some ICICI Excel statements are exported page-by-page into one worksheet.
    Later pages repeat the header but the physical Excel columns can shift
    left/right. Parsing only the first header therefore reads only page 1.
    """
    limit = len(raw) if max_rows is None else min(len(raw), max_rows)
    headers = []
    for i in range(limit):
        vals = raw.iloc[i].tolist()
        if is_generic_bank_header(vals):
            headers.append(i)

    # Avoid accidental adjacent duplicate header detections.
    cleaned = []
    for i in headers:
        if not cleaned or i - cleaned[-1] > 1:
            cleaned.append(i)
    return cleaned


def standardize_repeated_header_sheet(raw):
    """
    Parse a worksheet as independent header-delimited sections.

    Returns:
        tx_all          normalized transactions
        header_rows     detected header row indexes (0-based)
        candidate_rows  total rows presented to section standardizers

    Source Row is temporarily stored as _ABS_SOURCE_ROW so the caller can
    preserve the true worksheet row even when each page uses shifted columns.
    """
    header_rows = locate_all_generic_headers(raw)
    if not header_rows:
        return pd.DataFrame(), [], 0

    frames = []
    candidate_rows = 0

    for pos, header_row in enumerate(header_rows):
        next_header = header_rows[pos + 1] if pos + 1 < len(header_rows) else len(raw)
        headers = [clean_text(x) for x in raw.iloc[header_row].tolist()]

        safe_headers = []
        used = {}
        for col_no, h in enumerate(headers):
            base = h or f"Column_{col_no+1}"
            used[base] = used.get(base, 0) + 1
            safe_headers.append(base if used[base] == 1 else f"{base}_{used[base]}")

        section = raw.iloc[header_row + 1:next_header].copy()
        section.columns = safe_headers

        blank = section.apply(
            lambda row: all(clean_text(v) == "" for v in row.tolist()),
            axis=1,
        )
        section = section.loc[~blank].copy()
        candidate_rows += len(section)
        if section.empty:
            continue

        tx = standardize_generic_table(section)
        if tx.empty:
            continue

        # Pandas index is still the original raw worksheet index.
        tx["_ABS_SOURCE_ROW"] = tx.index.to_series().astype(int) + 1
        tx["_HEADER_ROW"] = header_row + 1
        frames.append(tx)

    if not frames:
        return pd.DataFrame(), header_rows, candidate_rows

    out = pd.concat(frames, ignore_index=True)

    # Exact same source row should never appear twice. This is only a parser
    # overlap guard; it does NOT deduplicate genuine bank transactions.
    if "_ABS_SOURCE_ROW" in out.columns:
        out = out.drop_duplicates(subset=["_ABS_SOURCE_ROW"], keep="first")

    return out, header_rows, candidate_rows


def dataframe_from_header(raw, header_row):
    headers = [clean_text(x) for x in raw.iloc[header_row].tolist()]
    safe_headers = []
    used = {}
    for i, h in enumerate(headers):
        base = h or f"Column_{i+1}"
        used[base] = used.get(base, 0) + 1
        if used[base] == 1:
            safe_headers.append(base)
        else:
            safe_headers.append(f"{base}_{used[base]}")
    data = raw.iloc[header_row + 1:].copy()
    data.columns = safe_headers
    blank = data.apply(
        lambda row: all(clean_text(v) == "" for v in row.tolist()),
        axis=1,
    )
    return data.loc[~blank].copy()


# ============================================================
# CUSTOMER / UTR EXTRACTORS  (expanded for all banks)
# ============================================================
LOCATION_WORDS = [
    "KILAKARAI", "BANGALORE", "DOMLUR", "RPC MUMBAI",
    "RPC-NASIK", "CHENNAI RPC", "RPC JODHPUR",
    "RPC-CHH", "SAMBHAJINAGAR", "DELHI", "MUMBAI",
    "KOLKATA", "HYDERABAD", "PUNE", "AHMEDABAD",
    "JAIPUR", "LUCKNOW", "KANPUR", "NAGPUR",
    "INDORE", "THANE", "BHOPAL", "VISAKHAPATNAM",
    "PATNA", "VADODARA", "GHAZIABAD", "LUDHIANA",
    "AGRA", "NASHIK", "RANCHI", "MEERUT",
    "RAJKOT", "VARANASI", "SRINAGAR", "AURANGABAD",
    "DHANBAD", "AMRITSAR", "NAVI MUMBAI",
    "ALLAHABAD", "RANCHI", "COIMBATORE", "MADURAI",
    "TRICHY", "SALEM", "TIRUNELVELI", "ERODE",
    "VELLORE", "TUTICORIN", "THIRUVANANTHAPURAM",
    "KOCHI", "KOZHIKODE", "MANGALORE", "MYSORE",
    "HUBLI", "BELGAUM", "GULBARGA", "DAVANGERE",
    "BELLARY", "SHIMOGA", "TUMKUR", "RAICHUR",
    "BIJAPUR", "UDUPI", "HASSAN", "MANDYA",
]


def clean_customer_name(value):
    s = clean_text(value).strip(" -/")
    for loc in LOCATION_WORDS:
        s = re.sub(
            r"\b" + re.escape(loc) + r"\b.*$",
            "",
            s,
            flags=re.I,
        ).strip(" -/")
    return s[:120]


def is_plain_customer_line(line):
    s = clean_text(line)
    if not s or len(s) > 90:
        return False
    reject_starts = [
        "total", "page ", "cont", "sr245",
        "this is an authenticated",
        "tran date", "your details",
        "your base branch", "summary of accounts",
        "statement of transactions", "regd address",
        "legend for transactions", "sincerely",
        "team icici", "cin :", "corporate office",
        "opening balance", "closing balance",
        "brought forward", "carried forward",
        "date", "particulars", "narration",
        "description", "cheque", "reference",
        "debit", "credit", "balance",
        "withdrawal", "deposit", "amount",
        "account no", "account number",
        "branch", "ifsc", "micr",
        "customer id", "customer name",
        "statement period", "from", "to",
        "page no", "printed on",
    ]
    if any(s.lower().startswith(x) for x in reject_starts):
        return False
    if "/" in s or ":" in s:
        return False
    digits = sum(ch.isdigit() for ch in s)
    letters = sum(ch.isalpha() for ch in s)
    if digits >= 3:
        return False
    return letters >= 4 and letters / max(len(s), 1) >= 0.50


def extract_reference_from_text(block):
    s = clean_text(block)
    patterns = [
        r"\bUPI/([0-9]{8,20})",
        r"\bMMT/IMPS/([0-9]{8,20})",
        r"\bINF/(?:NEFT|INFT)/([0-9]{8,20})",
        r"\bBIL/INFT/([0-9]{8,20})",
        r"\bNEFT-([A-Z0-9]{8,35})",
        r"\bRTGS-([A-Z0-9]{8,40})",
        r"\bRTGS/([A-Z0-9]{8,40})",
        r"\bTRF/[^/]+/([0-9]{3,20})",
        # HDFC patterns
        r"\bN\d{12,20}\b",
        r"\bIMPS\d{10,20}\b",
        # SBI patterns
        r"\bSBI[A-Z0-9]{10,25}\b",
        # Axis patterns
        r"\bAXIS[A-Z0-9]{8,25}\b",
        # Generic UTR (12-digit number)
        r"\bUTR\s*[:\-]?\s*(\d{12,16})\b",
        # Generic ref
        r"\bREF\s*[:\-]?\s*([A-Z0-9]{8,25})\b",
        # NEFT/RTGS generic
        r"\b(?:NEFT|RTGS)\s*[:\-/]\s*([A-Z0-9]{8,35})\b",
        # IMPS generic
        r"\bIMPS\s*[:\-/]\s*(\d{10,20})\b",
        # UPI ref
        r"\bUPI\s*[:\-/]\s*(\d{10,20})\b",
    ]
    for pattern in patterns:
        m = re.search(pattern, s, flags=re.I)
        if m:
            # Some bank-specific reference patterns intentionally match the
            # complete token and therefore have no capture group.  Older V7.3
            # always called group(1), which raised IndexError on those rows.
            # Use capture group 1 when present; otherwise use the full match.
            value = m.group(1) if m.lastindex else m.group(0)
            return clean_text(value)
    return ""


def extract_customer_from_text(block, prefix_name=""):
    if prefix_name:
        return clean_customer_name(prefix_name)
    s = clean_text(block)
    patterns = [
        # INF/NEFT/REF/IFSC/Name
        r"INF/NEFT/\d+/[A-Z0-9]+/"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,70}?)"
        r"(?=\s+(?:KILAKARAI|RPC|CHENNAI|BANGALORE|MUMBAI|DELHI)|$)",
        # RTGS/REF/IFSC/Name
        r"RTGS/[A-Z0-9]+/[A-Z0-9]+/"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,70}?)"
        r"(?=\s+(?:KILAKARAI|RPC|CHENNAI|BANGALORE|MUMBAI|DELHI)|$)",
        # IMPS/REF/Name/
        r"MMT/IMPS/\d+/([^/]{2,70})/",
        # BIL/INFT/ref/NA/ Name
        r"BIL/INFT/\d+/(?:NA|MIB-)/\s*"
        r"([A-Za-z][A-Za-z .&_-]{2,70}?)"
        r"(?=\s+(?:KILAKARAI|RPC|CHENNAI|BANGALORE|MUMBAI|DELHI)|$)",
        # NEFT-Nxxx-NAME
        r"NEFT-[A-Z0-9]+-"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,90}?)"
        r"(?=\s+(?:RPC|KILAKARAI|CHENNAI|BANGALORE|MUMBAI)|-SP\d|$)",
        # /NA/ NAME general
        r"/NA/\s*"
        r"([A-Za-z][A-Za-z .&_-]{2,70}?)"
        r"(?=\s+(?:KILAKARAI|RPC|CHENNAI|BANGALORE|MUMBAI|DELHI)|$)",
        # TO/ FROM name patterns (common in SBI, HDFC, Axis)
        r"\b(?:TO|FROM|A/C\s+OF|ACCT\s+OF|PAYMENT\s+TO|RECD\s+FROM)\s+"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,60}?)"
        r"(?=\s*(?:A/C|AC|ACC|IFSC|REF|UTR|NEFT|RTGS|IMPS|UPI|$))",
        # POS/ ECOM merchant name
        r"\b(?:POS|ECOM|ONLINE|SWIPE)\s+(?:AT\s+)?/?"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,50}?)"
        r"(?=\s*(?:ON|DT|DATE|REF|$))",
        # UPI-NAME@bank
        r"\bUPI[-/][A-Za-z0-9]+[-/]"
        r"([A-Za-z][A-Za-z0-9 .&_-]{2,40}?)"
        r"(?=\s*(?:@|REF|$))",
    ]
    for pattern in patterns:
        m = re.search(pattern, s, flags=re.I)
        if m:
            name = clean_customer_name(m.group(1))
            if name:
                return name
    return ""


# ============================================================
# GENERIC TABULAR STANDARDIZER
# ============================================================
def standardize_generic_table(table):
    date_col = find_column(table.columns, "date")
    narration_col = find_column(table.columns, "narration")
    debit_col = find_column(table.columns, "debit")
    credit_col = find_column(table.columns, "credit")
    balance_col = find_column(table.columns, "balance")
    customer_col = find_column(table.columns, "customer")
    utr_col = find_column(table.columns, "utr")
    amount_col = find_column(table.columns, "amount")
    type_col = find_column(table.columns, "type")

    if date_col is None:
        return pd.DataFrame()

    out = pd.DataFrame(index=table.index)
    out["Date"] = table[date_col].apply(parse_date_flexible)
    out["Narration"] = (
        table[narration_col] if narration_col else ""
    )
    out["Customer Name"] = (
        table[customer_col] if customer_col else ""
    )
    out["UTR / Reference"] = (
        table[utr_col] if utr_col else ""
    )
    out["Debit"] = (
        table[debit_col].apply(parse_number)
        if debit_col else np.nan
    )
    out["Credit"] = (
        table[credit_col].apply(parse_number)
        if credit_col else np.nan
    )
    out["Balance"] = (
        table[balance_col].apply(parse_number)
        if balance_col else np.nan
    )

    # Bank exports commonly use 0.00 placeholders in both Debit and Credit
    # columns. Treat zero as blank for direction purposes; otherwise every
    # credit row with Debit=0.00 is later misclassified as DEBIT.
    if debit_col is not None:
        out["Debit"] = pd.to_numeric(out["Debit"], errors="coerce").abs()
        out.loc[out["Debit"].abs() <= 10 ** (-(ROUND_DECIMALS + 1)), "Debit"] = np.nan
    if credit_col is not None:
        out["Credit"] = pd.to_numeric(out["Credit"], errors="coerce").abs()
        out.loc[out["Credit"].abs() <= 10 ** (-(ROUND_DECIMALS + 1)), "Credit"] = np.nan

    # Amount + DR/CR type fallback
    if (
        debit_col is None
        and credit_col is None
        and amount_col
        and type_col
    ):
        amounts = table[amount_col].apply(parse_number)
        txn_type = table[type_col].astype(str).str.upper()
        out["Debit"] = np.where(
            txn_type.str.contains(
                r"\bDR\b|DEBIT|WITHDRAW", regex=True, na=False
            ),
            amounts,
            np.nan,
        )
        out["Credit"] = np.where(
            txn_type.str.contains(
                r"\bCR\b|CREDIT|DEPOSIT", regex=True, na=False
            ),
            amounts,
            np.nan,
        )

    # Amount-column DR/CR marker fallback (e.g. "1,250.00 DR").
    if (
        debit_col is None
        and credit_col is None
        and amount_col
        and type_col is None
    ):
        raw_amount = table[amount_col].astype(str)
        amounts = table[amount_col].apply(parse_number).abs()
        dr_marker = raw_amount.str.contains(r"\b(?:DR|DEBIT)\b", case=False, regex=True, na=False)
        cr_marker = raw_amount.str.contains(r"\b(?:CR|CREDIT)\b", case=False, regex=True, na=False)
        if bool((dr_marker | cr_marker).any()):
            out["Debit"] = np.where(dr_marker, amounts, np.nan)
            out["Credit"] = np.where(cr_marker, amounts, np.nan)
        else:
            signed_amounts = table[amount_col].apply(parse_number)
            out["Debit"] = np.where(signed_amounts < 0, signed_amounts.abs(), np.nan)
            out["Credit"] = np.where(signed_amounts > 0, signed_amounts, np.nan)

    # Best-effort customer extraction from narration
    customer_blank = (
        out["Customer Name"]
        .fillna("")
        .astype(str)
        .str.strip()
        .eq("")
    )
    out.loc[customer_blank, "Customer Name"] = out.loc[
        customer_blank, "Narration"
    ].apply(lambda x: extract_customer_from_text(x, ""))

    utr_blank = (
        out["UTR / Reference"]
        .fillna("")
        .astype(str)
        .str.strip()
        .eq("")
    )
    out.loc[utr_blank, "UTR / Reference"] = out.loc[
        utr_blank, "Narration"
    ].apply(extract_reference_from_text)

    out["Customer Source"] = np.where(
        customer_col is not None,
        "SOURCE COLUMN",
        np.where(
            out["Customer Name"].fillna("").astype(str).str.strip().ne(""),
            "EXTRACTED FROM NARRATION",
            "NOT AVAILABLE",
        ),
    )
    out["UTR Source"] = np.where(
        utr_col is not None,
        "SOURCE COLUMN",
        np.where(
            out["UTR / Reference"].fillna("").astype(str).str.strip().ne(""),
            "EXTRACTED FROM NARRATION",
            "NOT AVAILABLE",
        ),
    )

    actual = out["Debit"].notna() | out["Credit"].notna()
    return out.loc[actual].copy()


# ============================================================
# EXCEL / CSV READER
# ============================================================
def _read_delimited_robust(path):
    """Read CSV/TXT exports with delimiter/preamble/encoding tolerance."""
    import csv
    last_error = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with open(path, "r", encoding=enc, errors="strict", newline="") as f:
                sample = f.read(65536)
                f.seek(0)
                # Prefer the delimiter that produces a stable multi-column structure.
                # This avoids commas inside Indian currency values (1,25,000.00)
                # being mistaken for CSV separators in pipe/tab exports.
                sample_lines = [ln for ln in sample.splitlines() if ln.strip()][:200]
                best = None
                for cand in ("|", "\t", ";", ","):
                    counts = [ln.count(cand) for ln in sample_lines]
                    positive = [c for c in counts if c > 0]
                    if not positive:
                        continue
                    from collections import Counter
                    mode_count, mode_freq = Counter(positive).most_common(1)[0]
                    coverage = mode_freq / max(len(sample_lines), 1)
                    score = (coverage, mode_count, mode_freq)
                    if best is None or score > best[0]:
                        best = (score, cand)
                delimiter = best[1] if best else ","
                rows = list(csv.reader(f, delimiter=delimiter))
            if not rows:
                return [("DELIMITED", pd.DataFrame())]
            width = max(len(r) for r in rows)
            padded = [r + [""] * (width - len(r)) for r in rows]
            return [("DELIMITED", pd.DataFrame(padded, dtype=str))]
        except Exception as exc:
            last_error = exc
    raise last_error


def read_excel_csv(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt"):
        return _read_delimited_robust(path)
    if ext in (".html", ".htm"):
        tables = pd.read_html(path, header=None, keep_default_na=False)
        return [(f"HTML_{i+1}", t.astype(str)) for i, t in enumerate(tables)]

    engine = "xlrd" if ext == ".xls" else "openpyxl"
    try:
        book = pd.ExcelFile(path, engine=engine)
        outputs = []
        for sheet in book.sheet_names:
            try:
                raw = pd.read_excel(book, sheet_name=sheet, header=None, dtype=str, keep_default_na=False)
                outputs.append((sheet, raw))
            except Exception as exc:
                logger.warning("Sheet '%s' skipped: %s", sheet, exc)
        if outputs:
            return outputs
    except Exception as excel_exc:
        # Many bank '.xls' downloads are actually HTML tables with an XLS extension.
        try:
            tables = pd.read_html(path, header=None, keep_default_na=False)
            if tables:
                logger.info("HTML-disguised Excel detected: %s", os.path.basename(path))
                return [(f"HTML_XLS_{i+1}", t.astype(str)) for i, t in enumerate(tables)]
        except Exception:
            pass
        raise excel_exc
    return []


# ============================================================
# UNIVERSAL PDF TEXT PARSER  (works for ALL banks)
# ============================================================
def is_bank_statement_pdf(text):
    """Broad bank-statement fingerprint; generic unknown-bank layouts are allowed in V7.2."""
    t = clean_text(text).lower()
    bank_keywords = [
        "bank", "statement", "account", "balance", "transaction", "debit", "credit",
        "deposit", "withdrawal", "particulars", "narration", "description", "cheque",
        "reference", "opening balance", "closing balance", "ifsc", "branch",
    ]
    score = sum(1 for kw in bank_keywords if kw in t)
    date_present = any(p.search(t) for p, _ in UNIVERSAL_DATE_PATTERNS)
    money_present = bool(re.search(r"\d[\d,]*\.\d{2}", t))
    if score >= 3:
        return True
    if UNKNOWN_BANK_GENERIC_PARSER and score >= 2 and date_present and money_present:
        return True
    return False


def signed_balance_from_line(line):
    """
    Extract signed running balance from a transaction line.
    Handles:  1,23,456.78 Cr | 499.99 Dr | 0.00 | -1,234.56
    """
    # Strip leading dates
    stripped = line
    for pat in DATE_LINE_PATTERNS:
        stripped = pat.sub("", stripped, count=1)
    # Also strip a second date (value date)
    for pat in DATE_LINE_PATTERNS:
        stripped = pat.sub("", stripped, count=1)

    nums = MONEY_TOKEN_RE.findall(stripped)
    if not nums:
        return None
    bal = parse_number(nums[-1])
    if pd.isna(bal):
        return None
    if re.search(r"\bDr\b\s*$", line, flags=re.I):
        return -abs(float(bal))
    if re.search(r"\bCr\b\s*$", line, flags=re.I):
        return abs(float(bal))
    if abs(float(bal)) <= 0.000001:
        return 0.0
    # If no Cr/Dr suffix, return as-is (will be validated by movement)
    return float(bal)


def pdf_amount_and_explicit_direction(line):
    """Return a source-backed PDF transaction amount and optional direction.

    The final three money-like tokens are treated as
    Debit/Withdrawal, Credit/Deposit, Balance only when all three look like
    formatted monetary columns. This intentionally prefers an unresolved
    direction over manufacturing one from narration numbers.

    A trailing ``Cr``/``Dr`` belongs to the running balance and is never used as
    transaction direction.
    """
    stripped = str(line or "")
    if ALLOW_LEADING_SERIAL_BEFORE_DATE:
        stripped = re.sub(
            r"^\s*\d{1,7}\s*(?:[|.:;-]\s*)?(?=\d{1,2}[-/.])",
            "",
            stripped,
            count=1,
        )

    for _ in range(2):
        candidate = stripped.lstrip()
        removed = False
        for pat in DATE_LINE_PATTERNS:
            updated = pat.sub("", candidate, count=1)
            if updated != candidate:
                stripped = updated
                removed = True
                break
        if not removed:
            break

    money_matches = list(MONEY_TOKEN_RE.finditer(stripped))
    if len(money_matches) < 2:
        return None, ""

    money_tokens = [match.group(0) for match in money_matches]
    formatted_money = re.compile(r"^(?:\d{1,3}(?:,\d{2,3})+|\d+)\.\d{2}$")

    if len(money_tokens) >= 3:
        tail = money_tokens[-3:]
        separators = [
            stripped[money_matches[-3].end():money_matches[-2].start()],
            stripped[money_matches[-2].end():money_matches[-1].start()],
        ]
        looks_like_three_money_columns = (
            all(formatted_money.fullmatch(token) for token in tail)
            and all(not re.search(r"[A-Za-z0-9]", sep) for sep in separators)
        )

        if looks_like_three_money_columns:
            debit_candidate = parse_number(tail[0])
            credit_candidate = parse_number(tail[1])
            if pd.notna(debit_candidate) and pd.notna(credit_candidate):
                debit_value = abs(float(debit_candidate))
                credit_value = abs(float(credit_candidate))
                epsilon = 10 ** (-(ROUND_DECIMALS + 1))

                if debit_value > epsilon and credit_value <= epsilon:
                    return debit_value, "DEBIT"
                if credit_value > epsilon and debit_value <= epsilon:
                    return credit_value, "CREDIT"

    amount = parse_number(money_tokens[-2])
    if pd.isna(amount):
        return None, ""
    return abs(float(amount)), ""


def transaction_amount_from_line(line):
    """Backward-compatible amount-only wrapper for the PDF parser."""
    amount, _ = pdf_amount_and_explicit_direction(line)
    return amount


def parse_universal_pdf_text(path):
    """
    Universal PDF text parser for ALL bank statements.

    Strategy:
    1. Extract text from each page.
    2. Find lines starting with dates (all formats).
    3. Group lines between dates into transaction blocks.
    4. Extract money values from each block.
    5. Identify amount and balance using heuristics.
    6. Determine debit/credit from balance movement.
    7. Extract customer name and UTR from narration.
    8. Reconcile against printed page totals where available.
    """
    import pdfplumber

    transactions = []
    page_recon = []
    previous_signed_balance = None

    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text(
                x_tolerance=1,
                y_tolerance=3,
            ) or ""
            lines = [
                line.strip()
                for line in page_text.splitlines()
                if line.strip()
            ]

            # Printed page totals (various formats)
            total_match = re.search(
                r"(?:Total|Page\s+Total|Grand\s+Total)\s*[:\-]?\s*"
                r"([\d,]+\.\d{2})\s+"
                r"([\d,]+\.\d{2})",
                page_text,
                flags=re.I,
            )
            # Some banks print: Total Debit: xxx  Total Credit: yyy
            alt_total_match = re.search(
                r"(?:Total\s+)?(?:Debit|Withdrawal)s?\s*[:\-]?\s*"
                r"([\d,]+\.\d{2}).*?"
                r"(?:Total\s+)?(?:Credit|Deposit)s?\s*[:\-]?\s*"
                r"([\d,]+\.\d{2})",
                page_text,
                flags=re.I | re.S,
            )
            if total_match:
                printed_debit = parse_number(total_match.group(1))
                printed_credit = parse_number(total_match.group(2))
            elif alt_total_match:
                printed_debit = parse_number(alt_total_match.group(1))
                printed_credit = parse_number(alt_total_match.group(2))
            else:
                printed_debit = np.nan
                printed_credit = np.nan

            # Find all date-starting lines
            transaction_indexes = []
            for i, line in enumerate(lines):
                if match_date_at_line_start(line):
                    transaction_indexes.append(i)

            page_rows = []
            for position, line_index in enumerate(transaction_indexes):
                line = lines[line_index]
                next_index = (
                    transaction_indexes[position + 1]
                    if position + 1 < len(transaction_indexes)
                    else len(lines)
                )

                block_lines = lines[line_index:next_index]

                # Stop at footer/total lines
                cleaned_block = []
                for bl in block_lines:
                    bl_lower = bl.lower().strip()
                    if bl_lower.startswith("total"):
                        break
                    if bl_lower.startswith("page ") and len(bl) < 30:
                        break
                    if bl_lower.startswith("contd"):
                        break
                    if bl_lower.startswith("e. & o. e."):
                        break
                    cleaned_block.append(bl)
                block_lines = cleaned_block
                if not block_lines:
                    continue

                # Check for customer name line before date
                prefix_name = ""
                if line_index > 0:
                    prev = lines[line_index - 1]
                    if is_plain_customer_line(prev):
                        prefix_name = prev

                # Find balance line in block
                balance_line = None
                for bl in block_lines:
                    if re.search(r"\b(?:Cr|Dr)\b\s*$", bl, flags=re.I):
                        balance_line = bl
                        break
                    # Zero balance without Cr/Dr
                    if match_date_at_line_start(bl):
                        stripped_bl = bl
                        for pat in DATE_LINE_PATTERNS:
                            stripped_bl = pat.sub("", stripped_bl, count=1)
                        for pat in DATE_LINE_PATTERNS:
                            stripped_bl = pat.sub("", stripped_bl, count=1)
                        zero_nums = MONEY_TOKEN_RE.findall(stripped_bl)
                        if (
                            len(zero_nums) >= 2
                            and abs(parse_number(zero_nums[-1]) or 0)
                            <= 0.000001
                        ):
                            balance_line = bl
                            break
                    # Lines with enough money tokens (amount + balance)
                    nums_in_bl = MONEY_TOKEN_RE.findall(bl)
                    if len(nums_in_bl) >= 2:
                        balance_line = bl
                        break

                if balance_line is None:
                    continue

                current_balance = signed_balance_from_line(balance_line)
                if current_balance is None:
                    continue

                amount, explicit_direction = pdf_amount_and_explicit_direction(balance_line)
                date_str = match_date_at_line_start(line)
                txn_date = parse_date_flexible(date_str)
                if pd.isna(txn_date):
                    continue

                is_bf = "B/F" in line.upper() or "BROUGHT FORWARD" in line.upper()

                block_for_parse = " ".join(
                    ([prefix_name] if prefix_name else [])
                    + block_lines
                )
                customer = extract_customer_from_text(
                    block_for_parse, prefix_name
                )
                reference = extract_reference_from_text(block_for_parse)

                debit = np.nan
                credit = np.nan
                balance_recon_status = "UNVERIFIED"
                balance_recon_difference = np.nan

                if (
                    not is_bf
                    and amount is not None
                    and previous_signed_balance is not None
                ):
                    balance_change = current_balance - previous_signed_balance
                    balance_recon_difference = abs(balance_change) - abs(amount)
                    tolerance = max(0.02, abs(amount) * 0.000001)
                    if abs(balance_recon_difference) <= tolerance:
                        balance_recon_status = "PASS"
                        if balance_change < 0:
                            debit = amount
                        elif balance_change > 0:
                            credit = amount
                    else:
                        # If the statement itself exposes a debit/credit amount column,
                        # retain that source-backed direction but keep reconciliation CHECK.
                        balance_recon_status = "CHECK"
                        if explicit_direction == "DEBIT":
                            debit = amount
                        elif explicit_direction == "CREDIT":
                            credit = amount
                elif not is_bf and amount is not None:
                    # First transaction has no prior balance for movement control.
                    # Only use direction when the PDF exposes an unambiguous
                    # withdrawal/deposit amount pair. A trailing Cr/Dr belongs to
                    # the running balance and must never be treated as transaction direction.
                    if explicit_direction == "DEBIT":
                        debit = amount
                        balance_recon_status = "EXPLICIT AMOUNT COLUMN"
                    elif explicit_direction == "CREDIT":
                        credit = amount
                        balance_recon_status = "EXPLICIT AMOUNT COLUMN"
                    else:
                        balance_recon_status = "UNRESOLVED"

                previous_signed_balance = current_balance

                if is_bf:
                    continue

                page_rows.append({
                    "Date": txn_date,
                    "Customer Name": customer,
                    "UTR / Reference": reference,
                    "Narration": clean_text(" ".join(block_lines)),
                    "Debit": debit,
                    "Credit": credit,
                    "Balance": current_balance,
                    "Balance Reconciliation Status": balance_recon_status,
                    "Balance Reconciliation Difference": balance_recon_difference,
                    "Customer Source": (
                        "PDF NAME LINE / NARRATION"
                        if customer else "NOT AVAILABLE"
                    ),
                    "UTR Source": (
                        "PDF NARRATION"
                        if reference else "NOT AVAILABLE"
                    ),
                    "Source Page": page_no,
                    "Parser": "UNIVERSAL PDF TEXT + BALANCE RECON",
                })

            # Page reconciliation
            if page_rows:
                page_df = pd.DataFrame(page_rows)
                parsed_debit = float(page_df["Debit"].fillna(0).sum())
                parsed_credit = float(page_df["Credit"].fillna(0).sum())
            else:
                parsed_debit = 0.0
                parsed_credit = 0.0

            debit_diff = (
                parsed_debit - printed_debit
                if pd.notna(printed_debit) else np.nan
            )
            credit_diff = (
                parsed_credit - printed_credit
                if pd.notna(printed_credit) else np.nan
            )
            if pd.notna(printed_debit) and pd.notna(printed_credit):
                page_status = (
                    "OK"
                    if abs(debit_diff) <= 0.02 and abs(credit_diff) <= 0.02
                    else "CHECK"
                )
            else:
                page_status = "NO PRINTED TOTAL"

            page_recon.append({
                "Page": page_no,
                "Parsed Debit": parsed_debit,
                "Printed Debit": printed_debit,
                "Debit Difference": debit_diff,
                "Parsed Credit": parsed_credit,
                "Printed Credit": printed_credit,
                "Credit Difference": credit_diff,
                "Status": page_status,
            })
            transactions.extend(page_rows)

    return pd.DataFrame(transactions), pd.DataFrame(page_recon)


# ============================================================
# GENERIC PDF TABLE READER
# ============================================================
def extract_pdf_tables(path):
    import pdfplumber
    outputs = []
    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            for table_no, table in enumerate(tables, start=1):
                if not table:
                    continue
                width = max(len(row or []) for row in table)
                rows = []
                for row in table:
                    row = list(row or [])
                    row += [""] * (width - len(row))
                    rows.append(row)
                outputs.append((
                    f"PDF Page {page_no} Table {table_no}",
                    pd.DataFrame(rows),
                    page_no,
                ))
    return outputs


def get_first_pdf_page_text(path):
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        if not pdf.pages:
            return ""
        return pdf.pages[0].extract_text(
            x_tolerance=1, y_tolerance=3
        ) or ""


# ============================================================
# OPTIONAL OCR FALLBACK
# ============================================================
def ocr_pdf_optional(path):
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception:
        return []
    try:
        images = convert_from_path(path, dpi=300)
    except Exception:
        return []
    outputs = []
    for page_no, image in enumerate(images, start=1):
        try:
            text_value = pytesseract.image_to_string(
                image, config="--psm 6"
            )
        except Exception:
            continue
        lines = []
        for x in text_value.splitlines():
            if not clean_text(x):
                continue
            # Preserve table structure when OCR emits aligned whitespace.
            parts = [clean_text(v) for v in re.split(r"\t+|\s{2,}", x.strip())]
            lines.append(parts if parts else [clean_text(x)])
        if lines:
            width = max(len(row) for row in lines)
            lines = [row + [""] * (width - len(row)) for row in lines]
            outputs.append((
                f"OCR Page {page_no}",
                pd.DataFrame(lines),
                page_no,
            ))
    return outputs



def nonblank_nunique(series):
    """Count unique nonblank values without pandas downcasting warnings."""
    s = series.astype("string").str.strip()
    s = s[(s.notna()) & (s != "")]
    return int(s.nunique())


# ============================================================
# V6 — ACCOUNT PROFILE / CLASSIFICATION / AUDIT HELPERS
# ============================================================
def detect_bank_from_filename(filename):
    # V7.3.2 verified source-identity overrides. These are statement files whose
    # owning bank is known from the statement itself; transaction counterparty
    # bank names must not change the owner bank.
    fn = clean_text(filename).lower()
    verified = {
        "nayeem soa": "Yes Bank",
        "sr283403308": "ICICI Bank",
        "axcess plus": "Standard Chartered Bank",
        "ret_8r69112820": "South Indian Bank",
        "idfcfirstbankstatement": "IDFC FIRST Bank",
    }
    for token, bank in verified.items():
        if token in fn:
            return bank
    """Fallback bank detection when statement metadata is sparse."""
    n = norm(os.path.basename(filename))
    aliases = {
        "ICICI Bank": ["icici"],
        "State Bank of India (SBI)": ["sbi", "state bank"],
        "HDFC Bank": ["hdfc"],
        "Axis Bank": ["axis"],
        "Kotak Mahindra Bank": ["kotak"],
        "Punjab National Bank (PNB)": ["pnb", "punjab national"],
        "Bank of Baroda (BOB)": ["bob", "bank of baroda"],
        "Canara Bank": ["canara"],
        "Union Bank of India": ["union bank"],
        "IDBI Bank": ["idbi"],
        "Yes Bank": ["yes bank", "yesbank"],
        "IndusInd Bank": ["indusind"],
        "Federal Bank": ["federal"],
        "RBL Bank": ["rbl"],
        "Bandhan Bank": ["bandhan"],
        "AU Small Finance Bank": ["au small", "aubank"],
        "Jana Small Finance Bank": ["jana"],
        "Suryoday Small Finance Bank": ["suryoday"],
        "Utkarsh Small Finance Bank": ["utkarsh"],
        "Karur Vysya Bank": ["karur vysya", "kvb"],
        "Indian Overseas Bank": ["iob", "indian overseas"],
        "Bank of India": ["bank of india", "boi"],
        "Standard Chartered Bank": ["standard chartered", "scb"],
        "HSBC": ["hsbc"],
        "Citibank N.A.": ["citibank", "citi"],
        "Deutsche Bank": ["deutsche"],
        "DBS Bank India": ["dbs", "digibank"],
        "Bank of America": ["bank of america", "bofa"],
        "Barclays Bank": ["barclays"],
        "BNP Paribas": ["bnp paribas"],
        "JPMorgan Chase Bank": ["jpmorgan", "jp morgan"],
        "MUFG Bank": ["mufg", "tokyo mitsubishi"],
        "Mizuho Bank": ["mizuho"],
        "Sumitomo Mitsui Banking Corporation": ["sumitomo mitsui", "smbc"],
        "Bank of Bahrain & Kuwait": ["bahrain kuwait", "bbk"],
        "Doha Bank": ["doha bank"],
        "Emirates NBD": ["emirates nbd", "enbd"],
        "First Abu Dhabi Bank": ["first abu dhabi", "fab bank"],
        "Mashreq Bank": ["mashreq"],
        "Qatar National Bank": ["qatar national", "qnb"],
        "Société Générale": ["societe generale", "société générale"],
        "Bank of China": ["bank of china"],
        "Industrial and Commercial Bank of China": ["industrial and commercial bank of china", "icbc"],
        "KEB Hana Bank": ["keb hana", "hana bank"],
        "Shinhan Bank": ["shinhan"],
        "Woori Bank": ["woori"],
    }
    aliases.update({
        "IDFC FIRST Bank": ["idfc first", "idfcfirst"],
        "Indian Bank": ["indian bank"],
        "Central Bank of India": ["central bank"],
        "UCO Bank": ["uco"],
        "Bank of Maharashtra": ["bank of maharashtra", "bom"],
        "Punjab & Sind Bank": ["punjab sind", "punjab and sind", "psb"],
        "City Union Bank": ["city union", "cub"],
        "Tamilnad Mercantile Bank": ["tamilnad mercantile", "tmb"],
        "South Indian Bank": ["south indian", "sib"],
        "Karnataka Bank": ["karnataka bank"],
        "DCB Bank": ["dcb"],
        "CSB Bank": ["csb", "catholic syrian"],
        "Dhanlaxmi Bank": ["dhanlaxmi"],
        "Jammu & Kashmir Bank": ["jammu kashmir", "j k bank", "jk bank"],
        "Ujjivan Small Finance Bank": ["ujjivan"],
        "Equitas Small Finance Bank": ["equitas"],
        "ESAF Small Finance Bank": ["esaf"],
        "Capital Small Finance Bank": ["capital sfb"],
        "Unity Small Finance Bank": ["unity sfb"],
        "Shivalik Small Finance Bank": ["shivalik"],
        "Airtel Payments Bank": ["airtel payments", "airtel bank"],
        "India Post Payments Bank": ["ippb", "india post payments"],
        "Fino Payments Bank": ["fino payments", "fino bank"],
        "Jio Payments Bank": ["jio payments", "jio bank"],
        "NSDL Payments Bank": ["nsdl payments", "nsdl bank"],
    })
    for bank, keys in aliases.items():
        if any(k in n for k in keys):
            return bank
    return "Unknown Bank"


def extract_account_profile_from_text(text_value, filename=""):
    """Best-effort account metadata extraction. Missing values remain blank."""
    raw = text_value or ""
    flat = clean_text(raw)
    filename_bank = detect_bank_from_filename(filename)
    bank = filename_bank if filename_bank != "Unknown Bank" else detect_bank_from_text(raw)

    def first_match(patterns):
        for pat in patterns:
            m = re.search(pat, raw, flags=re.I | re.M)
            if m:
                return clean_text(m.group(1))
        return ""

    account_number = first_match([
        r"(?:Account\s*(?:No\.?|Number)|A/C\s*(?:No\.?|Number))\s*[:\-]?\s*([A-Z0-9X*\-]{4,30})",
        r"(?:Account)\s*[:\-]?\s*([0-9X*]{6,30})",
    ])
    ifsc = first_match([
        r"(?:IFSC(?:\s*Code)?)\s*[:\-]?\s*([A-Z]{4}0[A-Z0-9]{6})",
    ])
    holder = first_match([
        r"(?:Customer\s*Name|Account\s*Holder(?:\s*Name)?|Name)\s*[:\-]\s*([^\n\r]{3,100})",
    ])
    period_from = first_match([
        r"(?:Statement\s*Period|Period)\s*(?:From)?\s*[:\-]?\s*(\d{1,2}[-/.][A-Za-z0-9]{2,9}[-/.][0-9]{2,4})",
        r"From\s*[:\-]?\s*(\d{1,2}[-/.][A-Za-z0-9]{2,9}[-/.][0-9]{2,4})",
    ])
    period_to = first_match([
        r"(?:Statement\s*Period|Period).*?(?:To)\s*[:\-]?\s*(\d{1,2}[-/.][A-Za-z0-9]{2,9}[-/.][0-9]{2,4})",
        r"To\s*[:\-]?\s*(\d{1,2}[-/.][A-Za-z0-9]{2,9}[-/.][0-9]{2,4})",
    ])

    masked = account_number[-4:] if len(account_number) >= 4 else account_number
    account_key = f"{bank} | {masked}" if masked else f"{bank} | {os.path.basename(filename)}"

    return {
        "Source File": os.path.basename(filename),
        "Bank Name": bank,
        "Bank Category": bank_category(bank),
        "Account Holder": holder,
        "Account Number": account_number,
        "Account Last 4": masked,
        "IFSC": ifsc,
        "Statement From": period_from,
        "Statement To": period_to,
        "Account Key": account_key,
    }


def build_account_profiles(paths):
    """Read only statement headers/first page for account metadata."""
    profiles = []
    for path in paths:
        base = os.path.basename(path)
        ext = os.path.splitext(path)[1].lower()
        profile_text = ""
        try:
            if ext == ".pdf":
                profile_text = get_first_pdf_page_text(path)
            else:
                sources = read_excel_csv(path)
                chunks = []
                for _, raw in sources[:3]:
                    if raw is not None and not raw.empty:
                        subset = raw.head(50).fillna("").astype(str)
                        chunks.append("\n".join(" | ".join(row) for row in subset.values.tolist()))
                profile_text = "\n".join(chunks)
        except Exception:
            profile_text = ""
        profiles.append(extract_account_profile_from_text(profile_text, path))
    return pd.DataFrame(profiles).drop_duplicates(subset=["Source File"], keep="first")


def classify_transaction_type(narration):
    s = clean_text(narration).upper()
    rules = [
        ("UPI", ["UPI/", "UPI-", "UPI "]),
        ("IMPS", ["IMPS", "MMT/IMPS"]),
        ("RTGS", ["RTGS"]),
        ("NEFT", ["NEFT"]),
        ("CHEQUE", ["CHEQUE", "CHQ", "CTS"]),
        ("ATM WITHDRAWAL", ["ATM", "CASH WDL", "CASH WITHDRAWAL"]),
        ("CASH DEPOSIT", ["CASH DEP", "CASH DEPOSIT"]),
        ("POS / CARD", ["POS", "ECOM", "CARD", "SWIPE"]),
        ("BANK CHARGES", ["CHARGE", "CHGS", "FEE", "COMMISSION", "PENAL"]),
        ("GST / TAX", ["GST", "CGST", "SGST", "IGST", "TAX"]),
        ("INTEREST", ["INTEREST", "INT.PD", "INT PAID", "INT RECEIVED"]),
        ("REFUND / REVERSAL", ["REFUND", "REVERSAL", "REVERSED", "RETURN"]),
        ("INTERNAL TRANSFER", ["INTERNAL TRANSFER", "OWN ACCOUNT", "SELF TRANSFER"]),
        ("BANK TRANSFER", ["INFT", "TRF", "TRANSFER"]),
        ("NACH / ECS", ["NACH", "ECS", "ACH"]),
    ]
    for category, keywords in rules:
        if any(k in s for k in keywords):
            return category
    return "OTHER"



def _contains_term(text, term):
    """Match an audit keyword as a token/phrase, not as a substring inside another word."""
    normalized = clean_text(text).upper()
    escaped = re.escape(clean_text(term).upper())
    escaped = escaped.replace(r"\ ", r"\s+")
    return re.search(r"(?<![A-Z0-9])" + escaped + r"(?![A-Z0-9])", normalized) is not None


def _contains_any_term(text, terms):
    return any(_contains_term(text, term) for term in terms)


def classify_payment_mode_v851(narration):
    """Classify payment mode conservatively from explicit narration evidence."""
    s = clean_text(narration).upper()

    if _contains_any_term(s, [
        "BILLPAY", "BILL PAY", "BILL PAYMENT", "BBPS", "BILLDESK",
        "UTILITY BILL", "BILL PAYMENT TXN",
    ]):
        return "BILL PAYMENT"
    if _contains_term(s, "RTGS"):
        return "RTGS"
    if _contains_term(s, "NEFT"):
        return "NEFT"
    if _contains_term(s, "IMPS") or "MMT/IMPS" in s:
        return "IMPS"
    if re.search(r"(?<![A-Z0-9])UPI(?=[/\-\s]|$)", s):
        return "UPI"
    if _contains_any_term(s, ["CASH DEP", "CASH DEPOSIT", "BY CASH"]):
        return "CASH DEPOSIT"
    if _contains_any_term(s, ["CASH WDL", "CASH WITHDRAWAL", "ATM WDL", "ATM WITHDRAWAL"]):
        return "CASH WITHDRAWAL"
    if _contains_any_term(s, ["CHEQUE", "CHQ", "CTS"]):
        return "CHEQUE"
    if _contains_any_term(s, ["POS", "ECOM", "CARD", "SWIPE"]):
        return "CARD / POS"
    if _contains_any_term(s, ["NACH", "ECS", "ACH"]):
        return "NACH / ECS"
    if _contains_any_term(s, ["INFT", "TRF", "TRANSFER"]):
        return "BANK TRANSFER"
    return "OTHER"


def classify_charge_type_v851(narration):
    """Classify only explicitly evidenced bank charges, fees, or taxes."""
    s = clean_text(narration).upper()

    bill = _contains_any_term(s, [
        "BILLPAY", "BILL PAY", "BILL PAYMENT", "BBPS", "BILLDESK", "UTILITY BILL",
    ])
    gst = _contains_any_term(s, ["GST", "CGST", "SGST", "IGST", "SERVICE TAX"])
    fee = _contains_any_term(s, [
        "CHARGE", "CHARGES", "CHGS", "CHG", "FEE", "FEES",
        "COMMISSION", "COMM", "COMMN", "PENAL", "PROCESSING FEE",
        "SERVICE CHARGE", "TRANSACTION CHARGE", "TXN CHARGE",
        "SMS CHARGE", "ATM CHARGE", "ANNUAL FEE",
        "IMPS CHARGE", "NEFT CHARGE", "RTGS CHARGE", "UPI CHARGE",
        "ACH CHARGE", "NACH CHARGE", "CHEQUE RETURN CHARGE",
        "RETURN CHARGE", "BOUNCE CHARGE",
    ])

    if bill and fee:
        return "BILL PAYMENT FEE"
    if gst and fee:
        return "GST ON BANK CHARGES"
    if gst:
        return "GST / TAX"
    if fee:
        return "BANK CHARGES / FEES"
    return ""


def add_v851_audit_columns(df):
    """
    Adds reporting-only audit columns. Original transaction data is preserved.
    Charge Amount uses the actual debit/credit amount of a charge row only.
    """
    g = df.copy()

    g["Payment Mode"] = g["Narration"].map(classify_payment_mode_v851)
    g["Charge Type"] = g["Narration"].map(classify_charge_type_v851)
    g["Is Charge / Fee"] = g["Charge Type"].ne("").map({True: "YES", False: "NO"})

    # Source-backed net charge: debit is a charge; credit is a reversal/refund.
    # This prevents charge reversals from inflating "Total Actual Charges / Fees".
    g["Charge Amount"] = 0.0
    charge_mask = g["Charge Type"].ne("")
    debit_amount = pd.to_numeric(g["Debit"], errors="coerce").fillna(0).abs()
    credit_amount = pd.to_numeric(g["Credit"], errors="coerce").fillna(0).abs()
    g.loc[charge_mask, "Charge Amount"] = (
        debit_amount.loc[charge_mask] - credit_amount.loc[charge_mask]
    )

    g["Bank Charges / Fees"] = 0.0
    mask = g["Charge Type"].eq("BANK CHARGES / FEES")
    g.loc[mask, "Bank Charges / Fees"] = g.loc[mask, "Charge Amount"]

    g["GST / Tax Charges"] = 0.0
    mask = g["Charge Type"].isin(["GST / TAX", "GST ON BANK CHARGES"])
    g.loc[mask, "GST / Tax Charges"] = g.loc[mask, "Charge Amount"]

    g["Bill Payment Fees"] = 0.0
    mask = g["Charge Type"].eq("BILL PAYMENT FEE")
    g.loc[mask, "Bill Payment Fees"] = g.loc[mask, "Charge Amount"]

    return g


def mode_summary_v851(grp):
    g = add_v851_audit_columns(grp)
    rows = []
    preferred = [
        "BILL PAYMENT", "NEFT", "IMPS", "UPI", "RTGS",
        "CASH DEPOSIT", "CASH WITHDRAWAL", "CHEQUE",
        "CARD / POS", "NACH / ECS", "BANK TRANSFER", "OTHER"
    ]
    for mode in preferred:
        x = g[g["Payment Mode"].eq(mode)]
        if x.empty:
            continue
        rows.append({
            "Payment Mode": mode,
            "Transaction Count": len(x),
            "Debit Count": int(x["Debit"].notna().sum()),
            "Total Debit": float(x["Debit"].fillna(0).sum()),
            "Credit Count": int(x["Credit"].notna().sum()),
            "Total Credit": float(x["Credit"].fillna(0).sum()),
            "Net Credit - Debit": float(x["Credit"].fillna(0).sum() - x["Debit"].fillna(0).sum()),
        })
    return pd.DataFrame(rows)


def charges_summary_v851(grp):
    g = add_v851_audit_columns(grp)
    x = g[g["Charge Type"].ne("")].copy()
    if x.empty:
        return pd.DataFrame(columns=[
            "Charge Type", "Transaction Count", "Net Charge Amount",
            "Total Debit Charges", "Total Credit/Reversal", "First Date", "Last Date"
        ])
    rows = []
    for charge_type, c in x.groupby("Charge Type", sort=True):
        rows.append({
            "Charge Type": charge_type,
            "Transaction Count": len(c),
            "Net Charge Amount": float(c["Charge Amount"].fillna(0).sum()),
            "Total Debit Charges": float(c["Debit"].fillna(0).sum()),
            "Total Credit/Reversal": float(c["Credit"].fillna(0).sum()),
            "First Date": c["Date"].dropna().min() if c["Date"].notna().any() else pd.NaT,
            "Last Date": c["Date"].dropna().max() if c["Date"].notna().any() else pd.NaT,
        })
    return pd.DataFrame(rows)


def charge_ledger_v851(grp):
    g = add_v851_audit_columns(grp)
    x = g[g["Charge Type"].ne("")].copy()
    cols = [c for c in [
        "Date","Customer Name","UTR / Reference","Narration",
        "Payment Mode","Charge Type","Charge Amount",
        "Bank Charges / Fees","GST / Tax Charges","Bill Payment Fees",
        "Debit","Credit","Balance","Direction",
        "Source Part","Source Row","Review Status"
    ] if c in x.columns]
    return x[cols].sort_values(["Date","Source Row"], kind="stable", na_position="last").reset_index(drop=True)



def explicit_gst_component(narration):
    """Extract GST amount only when narration explicitly states an amount."""
    s = clean_text(narration)
    patterns = [
        r"(?:GST|IGST|CGST|SGST)\s*(?:AMT|AMOUNT)?\s*[:\-]?\s*(?:RS\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
        r"(?:RS\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)\s*(?:GST|IGST|CGST|SGST)",
    ]
    for pat in patterns:
        m = re.search(pat, s, flags=re.I)
        if m:
            return parse_number(m.group(1))
    return np.nan


def calculate_confidence(row):
    score = 0
    reasons = []

    if pd.notna(row.get("Date")):
        score += 15
    else:
        reasons.append("Missing date")

    if pd.notna(row.get("Transaction Amount")):
        score += 20
    else:
        reasons.append("Missing amount")

    if clean_text(row.get("Direction")) in ("DEBIT", "CREDIT"):
        score += 15
    else:
        reasons.append("Direction unresolved")

    if clean_text(row.get("UTR / Reference")):
        score += 15
    else:
        reasons.append("Missing UTR/reference")

    if clean_text(row.get("Customer Name")):
        score += 10
    else:
        reasons.append("Missing customer")

    if pd.notna(row.get("Balance")):
        score += 10

    recon = clean_text(row.get("Balance Reconciliation Status")).upper()
    if recon == "PASS":
        score += 10
    elif recon == "CHECK":
        reasons.append("Balance movement mismatch")
    elif recon in ("UNRESOLVED", "UNVERIFIED"):
        reasons.append("Balance reconciliation unavailable")

    if clean_text(row.get("Possible Duplicate")) != "YES":
        score += 5
    else:
        reasons.append("Possible duplicate")

    return min(score, 100), "; ".join(dict.fromkeys(reasons))

def add_format_diagnostic(source_file, source_part, status, sample_text="", header_hits_value="", note=""):
    if not DIAGNOSTIC_MODE:
        return
    sample = clean_text(sample_text)[:1500]
    format_diagnostic_records.append({
        "Source File": source_file,
        "Source Part": source_part,
        "Status": status,
        "Detected Bank": detect_bank_from_text(sample_text) if sample_text else detect_bank_from_filename(source_file),
        "Header Hits": header_hits_value,
        "Sample / First Content": sample,
        "Note": note,
    })


def run_internal_regression_tests():
    """Small dependency-free regression suite for critical parser invariants."""
    failures = []

    def check(name, condition):
        if not bool(condition):
            failures.append(name)

    check("parse trailing minus", parse_number("1,234.50-") == -1234.50)
    check("canonical BOB", canonical_bank_name("Bank of Baroda") == "Bank of Baroda (BOB)")
    check("canonical YES", canonical_bank_name("YES Bank") == "Yes Bank")

    sample = pd.DataFrame({
        "Date": ["01/09/2026", "02/09/2026"],
        "Narration": ["TEST CREDIT", "TEST DEBIT"],
        "Debit": ["0.00", "250.00"],
        "Credit": ["500.00", "0.00"],
        "Balance": ["1500.00", "1250.00"],
    })
    std = standardize_generic_table(sample)
    check("zero debit becomes blank", pd.isna(std.iloc[0]["Debit"]) and std.iloc[0]["Credit"] == 500.0)
    check("zero credit becomes blank", std.iloc[1]["Debit"] == 250.0 and pd.isna(std.iloc[1]["Credit"]))

    amount, direction = pdf_amount_and_explicit_direction(
        "01/09/2026 TEST 1,000.00 0.00 9,000.00 Cr"
    )
    check("pdf explicit debit", amount == 1000.0 and direction == "DEBIT")
    amount, direction = pdf_amount_and_explicit_direction(
        "01/09/2026 TEST 0.00 1,000.00 10,000.00 Cr"
    )
    check("pdf explicit credit", amount == 1000.0 and direction == "CREDIT")
    amount, direction = pdf_amount_and_explicit_direction(
        "01/09/2026 TEST 1,000.00 10,000.00 Cr"
    )
    check("balance Cr is not txn direction", amount == 1000.0 and direction == "")

    amount, direction = pdf_amount_and_explicit_direction(
        "01/09/2026 REF 0 1,000.00 10,000.00 Cr"
    )
    check(
        "narration zero is not debit-credit column",
        amount == 1000.0 and direction == "",
    )

    ref = extract_reference_from_text("NEFT-ABCD12345678 PAYMENT")
    check("reference extraction", bool(ref))

    check("coffee is not fee", classify_charge_type_v851("COFFEE SHOP PURCHASE") == "")
    check("beach is not ACH", classify_payment_mode_v851("BEACH RESORT PAYMENT") == "OTHER")
    check("explicit fee detected", classify_charge_type_v851("ATM CHARGE 25.00") == "BANK CHARGES / FEES")

    charge_sample = pd.DataFrame({
        "Narration": ["ATM CHARGE", "ATM CHARGE REVERSAL"],
        "Debit": [25.0, np.nan],
        "Credit": [np.nan, 25.0],
    })
    charge_audit = add_v851_audit_columns(charge_sample)
    check(
        "charge reversal nets to zero",
        float(charge_audit["Charge Amount"].sum()) == 0.0
        and charge_audit.iloc[0]["Charge Amount"] == 25.0
        and charge_audit.iloc[1]["Charge Amount"] == -25.0,
    )

    if failures:
        raise AssertionError("Regression test failure(s): " + ", ".join(failures))
    print(f"SELF-TEST PASS: {14} critical checks")
    return True


if "--self-test" in sys.argv:
    run_internal_regression_tests()
    sys.exit(0)


# ============================================================
# LOAD SOURCES — V7.3.3 AUDIT / RAW / ROW-CONTROL
# ============================================================
all_frames = []
source_map = []
all_pdf_recon = []

logger.info("=" * 82)
logger.info("UNIVERSAL BANK BOOKS REPORT %s", VERSION)
logger.info("Author: %s", AUTHOR_NAME)
logger.info("Folder: %s", BASE_DIR)
logger.info("Source files found: %s", len(input_files))

for path in input_files:
    base = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()
    logger.info("Reading: %s", base)

    if ext == ".pdf":
        try:
            first_text = get_first_pdf_page_text(path)
        except Exception as exc:
            source_map.append({"Source File": base, "Source Part": "PDF", "Status": "READ ERROR", "Transaction Count": 0, "Note": str(exc)})
            log_exception("PDF Read Error", exc, base, "PDF")
            continue

        # V7.3.2: prefer a verified filename identity when one exists.
        # Otherwise inspect statement header text. This prevents counterparty
        # IFSC/bank names inside transactions from hijacking bank ownership.
        filename_bank = detect_bank_from_filename(base)
        detected_bank = filename_bank if filename_bank != "Unknown Bank" else detect_bank_from_text(first_text)
        logger.info("  Detected bank: %s", detected_bank)

        if not is_bank_statement_pdf(first_text):
            source_map.append({"Source File": base, "Source Part": "PDF", "Status": "NOT A BANK STATEMENT", "Transaction Count": 0, "Note": "PDF does not appear to be a bank statement."})
            log_exception("Not Bank Statement", "PDF did not meet bank-statement fingerprint threshold", base, "PDF")
            add_format_diagnostic(base, "PDF", "FINGERPRINT REVIEW", first_text, note="Bank-statement fingerprint was weak; inspect first-page text.")
            row_control_seed.append({"Source File": base, "Source Part": "PDF", "Source Rows": 0, "Candidate Data Rows": 0, "Imported Rows": 0, "Processed Rows": 0, "Rejected Rows": 0, "Output Rows": 0, "Status": "REVIEW", "Note": "Not recognized as a bank statement"})
            continue

        try:
            pdf_tx, pdf_recon = parse_universal_pdf_text(path)
        except Exception as exc:
            log_exception("PDF Parse Error", traceback.format_exc(), base, "UNIVERSAL PDF TEXT")
            logger.error("  Universal PDF parser failed: %s", exc)
            pdf_tx = pd.DataFrame()
            pdf_recon = pd.DataFrame()

        if not pdf_tx.empty:
            pdf_tx["Source File"] = base
            pdf_tx["Source Part"] = "Page " + pdf_tx["Source Page"].astype(str)
            pdf_tx["Source Row"] = ""
            all_frames.append(pdf_tx)

            if not pdf_recon.empty:
                pdf_recon["Source File"] = base
                all_pdf_recon.append(pdf_recon)
                bad_pages = int(pdf_recon["Status"].eq("CHECK").sum())
            else:
                bad_pages = 0

            if PRESERVE_RAW_DATA:
                store_raw_data(pdf_tx.copy(), base, "PDF_PARSED_TRANSACTIONS")

            source_map.append({
                "Source File": base,
                "Source Part": f"UNIVERSAL PDF ({detected_bank})",
                "Status": "RECOGNIZED" if bad_pages == 0 else "RECOGNIZED / RECON CHECK",
                "Transaction Count": len(pdf_tx),
                "Note": f"Page reconciliation issues: {bad_pages}",
            })
            row_control_seed.append({
                "Source File": base,
                "Source Part": "UNIVERSAL PDF TEXT",
                "Source Rows": len(pdf_tx),
                "Candidate Data Rows": len(pdf_tx),
                "Imported Rows": len(pdf_tx),
                "Processed Rows": len(pdf_tx),
                "Rejected Rows": 0,
                "Output Rows": len(pdf_tx),
                "Status": "PASS" if bad_pages == 0 else "REVIEW",
                "Note": f"PDF page reconciliation CHECK count: {bad_pages}",
            })
            parser_recon_records.append({
                "Source File": base,
                "Parser": "UNIVERSAL PDF TEXT + BALANCE RECON",
                "Transactions": len(pdf_tx),
                "Debit Total": float(pdf_tx["Debit"].fillna(0).sum()),
                "Credit Total": float(pdf_tx["Credit"].fillna(0).sum()),
                "PDF CHECK Pages": bad_pages,
                "Status": "PASS" if bad_pages == 0 else "REVIEW",
            })
            logger.info("  Universal PDF parser: %s transaction(s)", f"{len(pdf_tx):,}")
            continue

        logger.info("  Text parser found no transactions. Trying PDF table extraction / OCR.")
        try:
            pdf_sources = extract_pdf_tables(path)
        except Exception as exc:
            pdf_sources = []
            log_exception("PDF Table Extract Error", traceback.format_exc(), base, "PDF TABLE")
            logger.error("  PDF table extraction failed: %s", exc)

        if not pdf_sources and OCR_ENABLED:
            pdf_sources = ocr_pdf_optional(path)

        recognized_pdf = False
        for source_name, raw, page_no in pdf_sources:
            if PRESERVE_RAW_DATA:
                store_raw_data(raw.copy(), base, source_name)

            header_row = locate_generic_header(raw)
            if header_row is None:
                sample = " | ".join(clean_text(x) for x in raw.head(8).fillna("").astype(str).values.flatten().tolist())
                add_format_diagnostic(base, source_name, "NO HEADER", sample, note="PDF table/OCR extracted but no supported header was detected.")
                row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": 0, "Imported Rows": 0, "Processed Rows": 0, "Rejected Rows": 0, "Output Rows": 0, "Status": "REVIEW", "Note": "No bank header detected"})
                continue

            table = dataframe_from_header(raw, header_row)
            tx = standardize_generic_table(table)
            if tx.empty:
                row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": len(table), "Imported Rows": 0, "Processed Rows": 0, "Rejected Rows": len(table), "Output Rows": 0, "Status": "REVIEW", "Note": "Header detected but no transactions"})
                continue

            tx["Source File"] = base
            tx["Source Part"] = source_name
            tx["Source Row"] = (tx.index.to_series().astype(int) + header_row + 2).values
            tx["Source Page"] = page_no
            tx["Parser"] = "GENERIC PDF TABLE / OCR"
            all_frames.append(tx)
            recognized_pdf = True
            source_map.append({"Source File": base, "Source Part": source_name, "Status": "RECOGNIZED", "Transaction Count": len(tx), "Note": f"Bank: {detected_bank}; header row: {header_row + 1}"})
            row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": len(table), "Imported Rows": len(tx), "Processed Rows": len(tx), "Rejected Rows": max(len(table) - len(tx), 0), "Output Rows": len(tx), "Status": "PASS", "Note": "Generic PDF table/OCR parser"})
            parser_recon_records.append({"Source File": base, "Parser": "GENERIC PDF TABLE / OCR", "Transactions": len(tx), "Debit Total": float(tx["Debit"].fillna(0).sum()), "Credit Total": float(tx["Credit"].fillna(0).sum()), "PDF CHECK Pages": "", "Status": "PASS"})

        if not recognized_pdf:
            source_map.append({"Source File": base, "Source Part": "PDF", "Status": "UNRECOGNIZED", "Transaction Count": 0, "Note": f"Bank detected: {detected_bank}. No reliable transaction table recognized."})
            log_exception("Unrecognized PDF Format", "No reliable transaction rows recognized", base, "PDF")
            add_format_diagnostic(base, "PDF", "UNRECOGNIZED", first_text, note="Text, table and OCR parsers did not produce reliable transactions.")
        continue

    # EXCEL / CSV
    try:
        sources = read_excel_csv(path)
    except Exception as exc:
        source_map.append({"Source File": base, "Source Part": "", "Status": "READ ERROR", "Transaction Count": 0, "Note": str(exc)})
        log_exception("File Read Error", traceback.format_exc(), base, "")
        continue

    recognized_file = False
    for source_name, raw in sources:
        if PRESERVE_RAW_DATA:
            store_raw_data(raw.copy(), base, source_name)

        tx, header_rows, candidate_rows = standardize_repeated_header_sheet(raw)
        if not header_rows:
            sample = " | ".join(clean_text(x) for x in raw.head(8).fillna("").astype(str).values.flatten().tolist())
            add_format_diagnostic(base, source_name, "NO HEADER", sample, note="Excel/CSV/HTML/TXT source read successfully but no supported header was detected.")
            row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": 0, "Imported Rows": 0, "Processed Rows": 0, "Rejected Rows": 0, "Output Rows": 0, "Status": "REVIEW", "Note": "No bank header detected"})
            continue

        if tx.empty:
            row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": candidate_rows, "Imported Rows": 0, "Processed Rows": 0, "Rejected Rows": candidate_rows, "Output Rows": 0, "Status": "REVIEW", "Note": f"{len(header_rows)} header section(s) detected but no transactions"})
            continue

        tx["Source File"] = base
        tx["Source Part"] = source_name
        if "_ABS_SOURCE_ROW" in tx.columns:
            tx["Source Row"] = tx["_ABS_SOURCE_ROW"].astype(int)
            tx = tx.drop(columns=["_ABS_SOURCE_ROW"], errors="ignore")
        else:
            tx["Source Row"] = ""
        tx["Source Page"] = ""
        tx["Parser"] = "GENERIC EXCEL/CSV MULTI-HEADER"
        tx = tx.drop(columns=["_HEADER_ROW"], errors="ignore")
        all_frames.append(tx)
        recognized_file = True

        header_note = ", ".join(str(x + 1) for x in header_rows[:12])
        if len(header_rows) > 12:
            header_note += ", ..."
        source_map.append({"Source File": base, "Source Part": source_name, "Status": "RECOGNIZED", "Transaction Count": len(tx), "Note": f"{len(header_rows)} header section(s); header rows: {header_note}"})
        row_control_seed.append({"Source File": base, "Source Part": source_name, "Source Rows": len(raw), "Candidate Data Rows": candidate_rows, "Imported Rows": len(tx), "Processed Rows": len(tx), "Rejected Rows": max(candidate_rows - len(tx), 0), "Output Rows": len(tx), "Status": "PASS", "Note": f"Multi-header parser; {len(header_rows)} section(s)"})
        parser_recon_records.append({"Source File": base, "Parser": "GENERIC EXCEL/CSV MULTI-HEADER", "Transactions": len(tx), "Debit Total": float(tx["Debit"].fillna(0).sum()), "Credit Total": float(tx["Credit"].fillna(0).sum()), "PDF CHECK Pages": "", "Status": "PASS"})
        logger.info("  %s: %s transaction(s) across %s header section(s)", source_name, f"{len(tx):,}", len(header_rows))

    if not recognized_file:
        source_map.append({"Source File": base, "Source Part": "", "Status": "UNRECOGNIZED", "Transaction Count": 0, "Note": "Expected Date + Debit/Credit, or Date + Amount + DR/CR."})
        log_exception("Unrecognized Format", "No supported bank header found", base, "")
        add_format_diagnostic(base, "", "UNRECOGNIZED", note="No supported header found in any source part.")

if not all_frames:
    raise ValueError(
        "\\nNo bank transactions were recognized.\\n"
        "Expected Date + Debit/Credit or Date + Amount + DR/CR.\\n"
        "Searchable/text PDFs are preferred; OCR is optional for scanned PDFs."
    )

# ============================================================
# CONSOLIDATE
# ============================================================
transactions = pd.concat(all_frames, ignore_index=True, sort=False)

# ============================================================
# V8 SIMPLE — CONSOLIDATE, VALIDATE, SUMMARIZE, EXPORT
# ============================================================
# Keep the powerful parsers above. Everything below is intentionally small:
# one workbook, no draft ledgers, no PG fee modelling, no chargeback workflow,
# no RAW sheets by default, and no duplicate technical workbook.

for required_col in [
    "Date", "Customer Name", "UTR / Reference", "Narration",
    "Debit", "Credit", "Balance", "Customer Source", "UTR Source",
    "Source File", "Source Part", "Source Row", "Source Page", "Parser",
    "Balance Reconciliation Status", "Balance Reconciliation Difference",
]:
    if required_col not in transactions.columns:
        transactions[required_col] = ""

transactions["Date"] = pd.to_datetime(transactions["Date"], errors="coerce").dt.normalize()
for c in ["Debit", "Credit", "Balance", "Balance Reconciliation Difference"]:
    transactions[c] = pd.to_numeric(transactions[c], errors="coerce")

eps = 10 ** (-(ROUND_DECIMALS + 1))
has_debit = transactions["Debit"].notna() & transactions["Debit"].abs().gt(eps)
has_credit = transactions["Credit"].notna() & transactions["Credit"].abs().gt(eps)

# Never silently accept rows containing both debit and credit.
transactions["Direction"] = np.select(
    [has_debit & ~has_credit, has_credit & ~has_debit],
    ["DEBIT", "CREDIT"],
    default=""
)
transactions["Transaction Amount"] = np.select(
    [transactions["Direction"].eq("DEBIT"), transactions["Direction"].eq("CREDIT")],
    [transactions["Debit"], transactions["Credit"]],
    default=np.nan
)

# Account profile enrichment.
account_profile = build_account_profiles(input_files)
profile_lookup = (
    account_profile.set_index("Source File").to_dict("index")
    if not account_profile.empty else {}
)
for col in ["Bank Name", "Bank Category", "Account Holder", "Account Number",
            "Account Last 4", "IFSC", "Account Key"]:
    transactions[col] = transactions["Source File"].map(
        lambda f: profile_lookup.get(f, {}).get(col, "")
    )

transactions["Transaction Type"] = transactions["Narration"].apply(classify_transaction_type)
transactions["Explicit GST Component"] = transactions["Narration"].apply(explicit_gst_component)

# Stronger review controls.
def _simple_review(row):
    reasons = []
    if pd.isna(row.get("Date")):
        reasons.append("MISSING DATE")
    if pd.isna(row.get("Transaction Amount")):
        if has_debit.loc[row.name] and has_credit.loc[row.name]:
            reasons.append("BOTH DEBIT AND CREDIT")
        else:
            reasons.append("MISSING AMOUNT/DIRECTION")
    if clean_text(row.get("Balance Reconciliation Status")).upper() == "CHECK":
        reasons.append("BALANCE RECON CHECK")
    if not clean_text(row.get("UTR / Reference")):
        reasons.append("MISSING UTR")
    return " | ".join(reasons)

transactions["Review Reason"] = transactions.apply(_simple_review, axis=1)
transactions["Review Status"] = np.where(
    transactions["Review Reason"].eq(""), "OK", "REVIEW"
)

# Duplicate flag: conservative; do not delete anything.
dup_key = (
    transactions["Date"].astype(str) + "|" +
    transactions["Debit"].fillna(0).round(ROUND_DECIMALS).astype(str) + "|" +
    transactions["Credit"].fillna(0).round(ROUND_DECIMALS).astype(str) + "|" +
    transactions["UTR / Reference"].fillna("").astype(str).str.strip() + "|" +
    transactions["Narration"].fillna("").astype(str).str.strip()
)
transactions["Possible Duplicate"] = np.where(dup_key.duplicated(keep=False), "YES", "NO")
transactions.loc[transactions["Possible Duplicate"].eq("YES"), "Review Status"] = "REVIEW"
transactions.loc[
    transactions["Possible Duplicate"].eq("YES") & transactions["Review Reason"].eq(""),
    "Review Reason"
] = "POSSIBLE DUPLICATE"

# Final register order.
register_columns = [
    "Date", "Bank Name", "Account Key", "Account Number",
    "Customer Name", "UTR / Reference", "Narration", "Transaction Type",
    "Debit", "Credit", "Balance", "Direction", "Transaction Amount",
    "Explicit GST Component", "Review Status", "Review Reason",
    "Possible Duplicate", "Customer Source", "UTR Source",
    "Balance Reconciliation Status", "Balance Reconciliation Difference",
    "Parser", "Source File", "Source Part", "Source Page", "Source Row",
]
for c in register_columns:
    if c not in transactions.columns:
        transactions[c] = ""
transactions = transactions[register_columns].sort_values(
    ["Date", "Source File", "Source Page", "Source Row"], na_position="last"
).reset_index(drop=True)

review_required = transactions[
    transactions["Review Status"].eq("REVIEW")
].copy()

# Summaries.
transactions["Month"] = transactions["Date"].dt.to_period("M").astype(str)

date_wise = transactions.groupby("Date", dropna=False).agg(
    Transaction_Count=("Transaction Amount", "count"),
    Debit_Count=("Debit", lambda s: int(s.notna().sum())),
    Debit_Total=("Debit", "sum"),
    Credit_Count=("Credit", lambda s: int(s.notna().sum())),
    Credit_Total=("Credit", "sum"),
).reset_index()
date_wise["Net_Credit_Minus_Debit"] = date_wise["Credit_Total"].fillna(0) - date_wise["Debit_Total"].fillna(0)

month_wise = transactions.groupby("Month", dropna=False).agg(
    Transaction_Count=("Transaction Amount", "count"),
    Debit_Total=("Debit", "sum"),
    Credit_Total=("Credit", "sum"),
).reset_index()
month_wise["Net_Credit_Minus_Debit"] = month_wise["Credit_Total"].fillna(0) - month_wise["Debit_Total"].fillna(0)

customer_base = transactions[transactions["Customer Name"].fillna("").astype(str).str.strip().ne("")]
if customer_base.empty:
    customer_wise = pd.DataFrame(columns=[
        "Customer Name", "Transaction_Count", "Debit_Total", "Credit_Total",
        "Net_Credit_Minus_Debit", "First_Date", "Last_Date"
    ])
else:
    customer_wise = customer_base.groupby("Customer Name", dropna=False).agg(
        Transaction_Count=("Transaction Amount", "count"),
        Debit_Total=("Debit", "sum"),
        Credit_Total=("Credit", "sum"),
        First_Date=("Date", "min"),
        Last_Date=("Date", "max"),
    ).reset_index()
    customer_wise["Net_Credit_Minus_Debit"] = (
        customer_wise["Credit_Total"].fillna(0) - customer_wise["Debit_Total"].fillna(0)
    )

type_summary = transactions.groupby("Transaction Type", dropna=False).agg(
    Transaction_Count=("Transaction Amount", "count"),
    Debit_Total=("Debit", "sum"),
    Credit_Total=("Credit", "sum"),
).reset_index()

# UTR duplicate/conflict control.
utr_base = transactions[transactions["UTR / Reference"].fillna("").astype(str).str.strip().ne("")].copy()
if utr_base.empty:
    utr_check = pd.DataFrame(columns=["UTR / Reference", "Occurrences", "Unique Amounts", "Status"])
else:
    utr_check = utr_base.groupby("UTR / Reference").agg(
        Occurrences=("UTR / Reference", "size"),
        Unique_Amounts=("Transaction Amount", lambda s: s.dropna().round(ROUND_DECIMALS).nunique()),
        First_Date=("Date", "min"),
        Last_Date=("Date", "max"),
    ).reset_index()
    utr_check["Status"] = np.select(
        [utr_check["Unique_Amounts"].gt(1), utr_check["Occurrences"].gt(1)],
        ["CONFLICT - SAME UTR DIFFERENT AMOUNT", "REPEATED UTR"],
        default="OK"
    )
    utr_check = utr_check[utr_check["Status"].ne("OK")].copy()

# Reconciliation is per SOURCE FILE + ACCOUNT KEY to avoid mixing overlapping statements.
recon_rows = []
for (source_file, account_key), grp in transactions.groupby(["Source File", "Account Key"], dropna=False):
    grp = grp.sort_values(["Date", "Source Page", "Source Row"], na_position="last")
    with_bal = grp[grp["Balance"].notna()].copy()
    opening = closing = np.nan
    if not with_bal.empty:
        first = with_bal.iloc[0]
        opening = float(first["Balance"]) + (0.0 if pd.isna(first["Debit"]) else float(first["Debit"])) - (0.0 if pd.isna(first["Credit"]) else float(first["Credit"]))
        closing = float(with_bal.iloc[-1]["Balance"])
    debit_total = float(grp["Debit"].fillna(0).sum())
    credit_total = float(grp["Credit"].fillna(0).sum())
    expected = opening + credit_total - debit_total if pd.notna(opening) else np.nan
    diff = closing - expected if pd.notna(closing) and pd.notna(expected) else np.nan
    status = "OK" if pd.notna(diff) and abs(diff) <= 0.02 else ("CHECK" if pd.notna(diff) else "BALANCE NOT AVAILABLE")
    recon_rows.append({
        "Source File": source_file, "Account Key": account_key,
        "Bank Name": clean_text(grp.iloc[0].get("Bank Name")),
        "First Date": grp["Date"].min(), "Last Date": grp["Date"].max(),
        "Derived Opening Balance": opening, "Total Debit": debit_total,
        "Total Credit": credit_total, "Expected Closing Balance": expected,
        "Actual Closing Balance": closing, "Difference": diff, "Status": status,
    })
bank_reconciliation = pd.DataFrame(recon_rows)

source_mapping = pd.DataFrame(source_map)
pdf_page_recon = pd.concat(all_pdf_recon, ignore_index=True) if all_pdf_recon else pd.DataFrame()

# Compact data quality.
data_quality = pd.DataFrame([
    ["Total Transactions", len(transactions)],
    ["Debit Transactions", int(transactions["Direction"].eq("DEBIT").sum())],
    ["Credit Transactions", int(transactions["Direction"].eq("CREDIT").sum())],
    ["Review Required", len(review_required)],
    ["Possible Duplicate Rows", int(transactions["Possible Duplicate"].eq("YES").sum())],
    ["Missing Customer Name", int(transactions["Customer Name"].fillna("").astype(str).str.strip().eq("").sum())],
    ["Missing UTR / Reference", int(transactions["UTR / Reference"].fillna("").astype(str).str.strip().eq("").sum())],
    ["UTR Duplicate / Conflict Groups", len(utr_check)],
    ["Reconciliation CHECK", int(bank_reconciliation["Status"].eq("CHECK").sum()) if not bank_reconciliation.empty else 0],
], columns=["Metric", "Count"])

total_debit = float(transactions["Debit"].fillna(0).sum())
total_credit = float(transactions["Credit"].fillna(0).sum())

cover = pd.DataFrame([
    ["Report", "Universal Bank Books Report — Simple & Powerful"],
    ["Version", DISPLAY_VERSION],
    ["Author", AUTHOR_NAME],
    ["Instagram", AUTHOR_INSTAGRAM],
    ["Source Files", len(input_files)],
    ["Transactions", len(transactions)],
    ["Total Debit", total_debit],
    ["Total Credit", total_credit],
    ["Review Required", len(review_required)],
    ["Possible Duplicate Rows", int(transactions["Possible Duplicate"].eq("YES").sum())],
    ["Generated On", datetime.now().strftime("%d-%m-%Y %H:%M:%S")],
], columns=["Metric", "Value"])

# ============================================================
# V8.2 — BANK-WISE SEPARATE FILES + REAL RAW + MONTH-WISE + INTERBANK
# ============================================================
# IMPORTANT:
#   * No bank/source statement is mixed with another bank/source workbook.
#   * Every output workbook contains its own MONTHLY_SUMMARY and EASY_STATEMENT.
#   * REAL_RAW_* sheets are copied from source rows before standardization.
#   * INTERBANK_MATCHES contains candidate cross-bank debit/credit pairs only.
#     It does not invent or automatically confirm an internal transfer.

BANK_OUTPUT_DIR = os.path.join(BASE_DIR, "BANK_WISE_REPORTS_V8_5_1_AUDIT_MODES_CHARGES")
os.makedirs(BANK_OUTPUT_DIR, exist_ok=True)


def safe_file_component(value, max_len=70):
    s = clean_text(value)
    s = re.sub(r'[^A-Za-z0-9._-]+', '_', s).strip('._-')
    return (s or 'BANK_STATEMENT')[:max_len]


def safe_sheet_name(value, used=None):
    s = re.sub(r'[\\/*?:\[\]]+', '_', clean_text(value)).strip(" '") or 'Sheet'
    s = s[:31]
    if used is None:
        return s
    base = s
    n = 2
    while s.upper() in used:
        suffix = f'_{n}'
        s = base[:31-len(suffix)] + suffix
        n += 1
    used.add(s.upper())
    return s


def month_summary_for_group(grp):
    g = grp.copy()
    g = g[g['Date'].notna()].sort_values(['Date','Source Row'], kind='stable')
    if g.empty:
        return pd.DataFrame(columns=[
            'Year-Month','Month','Transaction Count','Debit Count','Total Debit',
            'Credit Count','Total Credit','Net Credit - Debit','Month End Available Balance'
        ])
    g['Year-Month'] = g['Date'].dt.to_period('M').astype(str)
    rows=[]
    for ym, m in g.groupby('Year-Month', sort=True):
        m=m.sort_values(['Date','Source Row'], kind='stable')
        bal=m['Balance'].dropna()
        rows.append({
            'Year-Month': ym,
            'Month': pd.Period(ym, freq='M').strftime('%B %Y'),
            'Transaction Count': len(m),
            'Debit Count': int(m['Direction'].eq('DEBIT').sum()),
            'Total Debit': float(m['Debit'].fillna(0).sum()),
            'Credit Count': int(m['Direction'].eq('CREDIT').sum()),
            'Total Credit': float(m['Credit'].fillna(0).sum()),
            'Net Credit - Debit': float(m['Credit'].fillna(0).sum()-m['Debit'].fillna(0).sum()),
            'Month End Available Balance': float(bal.iloc[-1]) if not bal.empty else np.nan,
        })
    return pd.DataFrame(rows)



def _customer_display_series(df):
    """
    Reporting-only customer label.

    IMPORTANT:
    The original Customer Name is never overwritten.
    Blank/unextractable customers are classified as UNIDENTIFIED / REVIEW so
    their debit/credit remains visible in customer control totals.
    """
    s = df['Customer Name'].fillna('').astype(str).str.strip()
    return s.where(s.ne(''), 'UNIDENTIFIED / REVIEW')


def customer_wise_summary_for_group(grp):
    """Lifetime customer summary for one bank/source file."""
    g = grp.copy()
    if g.empty:
        return pd.DataFrame(columns=[
            'Customer Name','Transaction Count','Debit Count','Total Debit',
            'Credit Count','Total Credit','Net Credit - Debit',
            'First Transaction Date','Last Transaction Date',
            'Months Active','Review Transaction Count'
        ])

    g['Customer Display'] = _customer_display_series(g)

    rows = []
    for customer, c in g.groupby('Customer Display', dropna=False, sort=True):
        c = c.sort_values(['Date','Source Row'], kind='stable', na_position='last')
        valid_dates = c['Date'].dropna()
        months_active = (
            valid_dates.dt.to_period('M').astype(str).nunique()
            if not valid_dates.empty else 0
        )
        rows.append({
            'Customer Name': customer,
            'Transaction Count': len(c),
            'Debit Count': int(c['Debit'].notna().sum()),
            'Total Debit': float(c['Debit'].fillna(0).sum()),
            'Credit Count': int(c['Credit'].notna().sum()),
            'Total Credit': float(c['Credit'].fillna(0).sum()),
            'Net Credit - Debit': float(c['Credit'].fillna(0).sum() - c['Debit'].fillna(0).sum()),
            'First Transaction Date': valid_dates.min() if not valid_dates.empty else pd.NaT,
            'Last Transaction Date': valid_dates.max() if not valid_dates.empty else pd.NaT,
            'Months Active': int(months_active),
            'Review Transaction Count': int(c['Review Status'].fillna('').astype(str).str.upper().eq('REVIEW').sum())
                if 'Review Status' in c.columns else 0,
        })

    out = pd.DataFrame(rows)
    if not out.empty:
        # Largest monetary relationship first, unidentified/review kept visible.
        out['Total Movement'] = out['Total Debit'].fillna(0) + out['Total Credit'].fillna(0)
        out = out.sort_values(
            ['Total Movement','Transaction Count','Customer Name'],
            ascending=[False,False,True],
            kind='stable'
        ).drop(columns=['Total Movement']).reset_index(drop=True)
    return out


def customer_monthly_summary_for_group(grp):
    """
    Most important V8.5 sheet:
    Customer + month level Debit/Credit statement for one bank/source file.
    """
    g = grp.copy()
    if g.empty:
        return pd.DataFrame(columns=[
            'Customer Name','Year-Month','Month','Transaction Count',
            'Debit Count','Total Debit','Credit Count','Total Credit',
            'Net Credit - Debit','First Transaction Date',
            'Last Transaction Date','Review Transaction Count'
        ])

    g['Customer Display'] = _customer_display_series(g)
    g = g[g['Date'].notna()].copy()
    if g.empty:
        return pd.DataFrame(columns=[
            'Customer Name','Year-Month','Month','Transaction Count',
            'Debit Count','Total Debit','Credit Count','Total Credit',
            'Net Credit - Debit','First Transaction Date',
            'Last Transaction Date','Review Transaction Count'
        ])

    g['Year-Month'] = g['Date'].dt.to_period('M').astype(str)

    rows = []
    for (customer, ym), c in g.groupby(['Customer Display','Year-Month'], dropna=False, sort=True):
        c = c.sort_values(['Date','Source Row'], kind='stable')
        rows.append({
            'Customer Name': customer,
            'Year-Month': ym,
            'Month': pd.Period(ym, freq='M').strftime('%B %Y'),
            'Transaction Count': len(c),
            'Debit Count': int(c['Debit'].notna().sum()),
            'Total Debit': float(c['Debit'].fillna(0).sum()),
            'Credit Count': int(c['Credit'].notna().sum()),
            'Total Credit': float(c['Credit'].fillna(0).sum()),
            'Net Credit - Debit': float(c['Credit'].fillna(0).sum() - c['Debit'].fillna(0).sum()),
            'First Transaction Date': c['Date'].min(),
            'Last Transaction Date': c['Date'].max(),
            'Review Transaction Count': int(c['Review Status'].fillna('').astype(str).str.upper().eq('REVIEW').sum())
                if 'Review Status' in c.columns else 0,
        })

    out = pd.DataFrame(rows)
    return out.sort_values(
        ['Customer Name','Year-Month'],
        ascending=[True,True],
        kind='stable'
    ).reset_index(drop=True)


def customer_ledger_for_group(grp):
    """
    Easy transaction-level customer ledger.
    Sorted customer -> date -> source row.
    Keeps the original Customer Name plus a reporting label for blanks.
    """
    g = grp.copy()
    g['Customer Display'] = _customer_display_series(g)

    cols = [c for c in [
        'Customer Display','Customer Name','Date','UTR / Reference','Narration',
        'Debit','Credit','Balance','Direction','Transaction Amount',
        'Transaction Type','Possible Duplicate','Review Status',
        'Source Part','Source Row'
    ] if c in g.columns]

    g = g[cols].copy()
    if 'Customer Display' in g.columns:
        g = g.rename(columns={'Customer Display':'Customer Reporting Name'})
    sort_cols = [c for c in ['Customer Reporting Name','Date','Source Row'] if c in g.columns]
    if sort_cols:
        g = g.sort_values(sort_cols, kind='stable', na_position='last')
    return g.reset_index(drop=True)


def customer_control_for_group(grp, customer_summary, customer_monthly):
    """Control totals proving customer reporting reconciles back to the bank source."""
    source_debit = float(grp['Debit'].fillna(0).sum())
    source_credit = float(grp['Credit'].fillna(0).sum())

    cw_debit = float(customer_summary['Total Debit'].fillna(0).sum()) if not customer_summary.empty else 0.0
    cw_credit = float(customer_summary['Total Credit'].fillna(0).sum()) if not customer_summary.empty else 0.0

    cm_debit = float(customer_monthly['Total Debit'].fillna(0).sum()) if not customer_monthly.empty else 0.0
    cm_credit = float(customer_monthly['Total Credit'].fillna(0).sum()) if not customer_monthly.empty else 0.0

    blank_customer_rows = int(
        grp['Customer Name'].fillna('').astype(str).str.strip().eq('').sum()
    )

    return pd.DataFrame([
        ['Source Transaction Count', len(grp)],
        ['Source Total Debit', source_debit],
        ['Source Total Credit', source_credit],
        ['Customer Summary Total Debit', cw_debit],
        ['Customer Summary Total Credit', cw_credit],
        ['Customer Summary Debit Difference', round(cw_debit - source_debit, ROUND_DECIMALS)],
        ['Customer Summary Credit Difference', round(cw_credit - source_credit, ROUND_DECIMALS)],
        ['Customer Monthly Total Debit', cm_debit],
        ['Customer Monthly Total Credit', cm_credit],
        ['Customer Monthly Debit Difference', round(cm_debit - source_debit, ROUND_DECIMALS)],
        ['Customer Monthly Credit Difference', round(cm_credit - source_credit, ROUND_DECIMALS)],
        ['Blank / Unidentified Customer Rows', blank_customer_rows],
        ['Customer Summary Status',
         'PASS' if round(cw_debit-source_debit, ROUND_DECIMALS) == 0
                   and round(cw_credit-source_credit, ROUND_DECIMALS) == 0 else 'REVIEW'],
        ['Customer Monthly Status',
         'PASS' if round(cm_debit-source_debit, ROUND_DECIMALS) == 0
                   and round(cm_credit-source_credit, ROUND_DECIMALS) == 0 else 'REVIEW'],
    ], columns=['Control','Value'])



def build_interbank_matches(tx):
    """One-to-one candidate matching across different source statements.

    Strongest evidence first:
      1) exact amount + exact same nonblank UTR/reference across debit/credit
      2) exact amount + same date
      3) exact amount + date within 1 day and transfer-like narration
    Nothing is deleted or auto-posted. These are review candidates only.
    """
    t=tx.copy()
    t=t[t['Date'].notna() & t['Transaction Amount'].notna() & t['Direction'].isin(['DEBIT','CREDIT'])].copy()
    t['AmtKey']=t['Transaction Amount'].round(ROUND_DECIMALS)
    t['UTRKey']=t['UTR / Reference'].fillna('').astype(str).str.strip().str.upper()
    deb=t[t['Direction'].eq('DEBIT')].copy()
    cre=t[t['Direction'].eq('CREDIT')].copy()
    used_d=set(); used_c=set(); matches=[]

    def transfer_like(s):
        return bool(re.search(r'\b(?:NEFT|RTGS|IMPS|INFT|TRANSFER|TRF|OWN|SELF|FUND|UPI)\b', clean_text(s), re.I))

    candidates=[]
    credit_by_amount={k:v for k,v in cre.groupby('AmtKey')}
    for di,d in deb.iterrows():
        cg=credit_by_amount.get(d['AmtKey'])
        if cg is None:
            continue
        for ci,c in cg.iterrows():
            if clean_text(d['Source File']) == clean_text(c['Source File']):
                continue
            daydiff=abs((pd.Timestamp(c['Date'])-pd.Timestamp(d['Date'])).days)
            same_utr=bool(d['UTRKey']) and d['UTRKey']==c['UTRKey']
            same_date=daydiff==0
            transfer_hint=transfer_like(d.get('Narration')) or transfer_like(c.get('Narration'))
            if same_utr:
                score=100; basis='EXACT AMOUNT + SAME UTR/REFERENCE'
            elif same_date:
                score=80; basis='EXACT AMOUNT + SAME DATE'
            elif daydiff<=1 and transfer_hint:
                score=60; basis='EXACT AMOUNT + DATE WITHIN 1 DAY + TRANSFER NARRATION'
            else:
                continue
            candidates.append((score,-daydiff,di,ci,basis,daydiff))

    # Highest-confidence one-to-one matching only.
    candidates.sort(reverse=True)
    for score, negdiff, di, ci, basis, daydiff in candidates:
        if di in used_d or ci in used_c:
            continue
        used_d.add(di); used_c.add(ci)
        d=deb.loc[di]; c=cre.loc[ci]
        matches.append({
            'Match Status': 'CANDIDATE - VERIFY',
            'Confidence': 'HIGH' if score>=100 else ('MEDIUM' if score>=80 else 'REVIEW'),
            'Match Basis': basis,
            'Amount': float(d['Transaction Amount']),
            'Debit Date': d['Date'],
            'Debit Bank': d.get('Bank Name',''),
            'Debit Account': d.get('Account Number',''),
            'Debit UTR / Ref': d.get('UTR / Reference',''),
            'Debit Narration': d.get('Narration',''),
            'Debit Source File': d.get('Source File',''),
            'Credit Date': c['Date'],
            'Credit Bank': c.get('Bank Name',''),
            'Credit Account': c.get('Account Number',''),
            'Credit UTR / Ref': c.get('UTR / Reference',''),
            'Credit Narration': c.get('Narration',''),
            'Credit Source File': c.get('Source File',''),
            'Date Difference Days': daydiff,
        })
    return pd.DataFrame(matches)


interbank_matches = build_interbank_matches(transactions)

# Map raw captures to source file. raw_data_sheets contains the actual raw DataFrames
# produced by read_excel_csv()/PDF table extraction before standardization.
raw_by_source = {}
for rec in raw_data_index_records:
    source = rec.get('Source File','')
    raw_name = rec.get('Raw Sheet','')
    part = rec.get('Source Part','')
    if raw_name in raw_data_sheets:
        raw_by_source.setdefault(source, []).append((part, raw_name, raw_data_sheets[raw_name]))

# If config.yaml from an older run disabled raw preservation, warn clearly.
if not PRESERVE_RAW_DATA:
    logger.warning('preserve_raw_data is FALSE in config.yaml. REAL RAW sheets cannot be created. Set preserve_raw_data: true.')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL_SIMPLE = PatternFill('solid', fgColor='1F4E78')
SUB_FILL_SIMPLE = PatternFill('solid', fgColor='D9EAF7')
WARN_FILL_SIMPLE = PatternFill('solid', fgColor='FFF2CC')
WHITE_BOLD_SIMPLE = Font(color='FFFFFF', bold=True)
BOLD_SIMPLE = Font(bold=True)
THIN_SIMPLE = Side(style='thin', color='D9D9D9')
BORDER_SIMPLE = Border(left=THIN_SIMPLE,right=THIN_SIMPLE,top=THIN_SIMPLE,bottom=THIN_SIMPLE)
MONEY_FMT_SIMPLE = '₹#,##0.00;[Red](₹#,##0.00);-'

MODE_FILLS_V851 = {
    "BILL PAYMENT": PatternFill("solid", fgColor="FFF2CC"),
    "NEFT": PatternFill("solid", fgColor="D9EAD3"),
    "IMPS": PatternFill("solid", fgColor="D9EAF7"),
    "UPI": PatternFill("solid", fgColor="EADCF8"),
    "RTGS": PatternFill("solid", fgColor="FCE5CD"),
    "CASH DEPOSIT": PatternFill("solid", fgColor="D0E0E3"),
    "CASH WITHDRAWAL": PatternFill("solid", fgColor="F4CCCC"),
    "CHEQUE": PatternFill("solid", fgColor="CFE2F3"),
    "CARD / POS": PatternFill("solid", fgColor="E2F0D9"),
    "NACH / ECS": PatternFill("solid", fgColor="D9D2E9"),
    "BANK TRANSFER": PatternFill("solid", fgColor="EDEDED"),
    "OTHER": PatternFill("solid", fgColor="FFFFFF"),
}
CHARGE_FILL_V851 = PatternFill("solid", fgColor="FFD966")
GST_FILL_V851 = PatternFill("solid", fgColor="F4B183")
BILL_FEE_FILL_V851 = PatternFill("solid", fgColor="FFE699")


def _safe(v):
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v,pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v,np.generic):
        return v.item()
    return v


def add_df_sheet(wb,name,df,money_cols=(),date_cols=(),highlight_review=False,highlight_modes=False,highlight_charges=False):
    used={x.upper() for x in wb.sheetnames}
    name=safe_sheet_name(name,used)
    ws=wb.create_sheet(name)
    if df is None or df.empty:
        ws['A1']='No data'
        return ws
    cols=list(df.columns)
    for c,col in enumerate(cols,1):
        cell=ws.cell(1,c,col)
        cell.fill=HEADER_FILL_SIMPLE; cell.font=WHITE_BOLD_SIMPLE
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.border=BORDER_SIMPLE
    for r,row in enumerate(df.itertuples(index=False,name=None),2):
        for c,val in enumerate(row,1):
            cell=ws.cell(r,c,_safe(val)); cell.border=BORDER_SIMPLE
            if cols[c-1] in money_cols and isinstance(cell.value,(int,float)):
                cell.number_format=MONEY_FMT_SIMPLE
            if cols[c-1] in date_cols and cell.value is not None:
                cell.number_format='dd-mm-yyyy'
        if highlight_review and 'Match Status' in cols:
            for cell in ws[r]:
                cell.fill=WARN_FILL_SIMPLE

        if highlight_modes and 'Payment Mode' in cols:
            mode = clean_text(df.iloc[r-2].get('Payment Mode'))
            fill = MODE_FILLS_V851.get(mode)
            if fill:
                for cell in ws[r]:
                    cell.fill = fill

        if highlight_charges and 'Charge Type' in cols:
            charge_type = clean_text(df.iloc[r-2].get('Charge Type'))
            if charge_type:
                if charge_type == 'BILL PAYMENT FEE':
                    fill = BILL_FEE_FILL_V851
                elif charge_type in ('GST / TAX', 'GST ON BANK CHARGES'):
                    fill = GST_FILL_V851
                else:
                    fill = CHARGE_FILL_V851
                for cell in ws[r]:
                    cell.fill = fill
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for idx,col in enumerate(cols,1):
        vals=[str(col)] + [str(x) for x in df[col].head(150).fillna('').tolist()]
        width=min(max(max(len(x) for x in vals)+2,10),45)
        ws.column_dimensions[get_column_letter(idx)].width=width
    return ws


def add_raw_sheet(wb,name,raw_df):
    """Write the raw source matrix exactly: no artificial DataFrame header row."""
    used={x.upper() for x in wb.sheetnames}
    ws=wb.create_sheet(safe_sheet_name(name,used))
    if raw_df is None or raw_df.empty:
        ws['A1']='No raw data captured'
        return ws
    for r_idx,row in enumerate(raw_df.itertuples(index=False,name=None),1):
        for c_idx,val in enumerate(row,1):
            ws.cell(r_idx,c_idx,_safe(val))
    ws.freeze_panes='A1'
    # Conservative widths; raw layout should remain readable, not reformatted into a new statement.
    max_cols=min(raw_df.shape[1],80)
    for c_idx in range(1,max_cols+1):
        sample=[]
        for v in raw_df.iloc[:120,c_idx-1].tolist():
            sample.append('' if v is None else str(v))
        width=min(max([len(x) for x in sample]+[8])+2,42)
        ws.column_dimensions[get_column_letter(c_idx)].width=width
    return ws


generated_files=[]
for seq,(source_file,grp) in enumerate(transactions.groupby('Source File',dropna=False,sort=False),1):
    grp=grp.sort_values(['Date','Source Row'],kind='stable',na_position='last').copy()
    first=grp.iloc[0]
    bank_name=clean_text(first.get('Bank Name')) or 'UNKNOWN_BANK'
    account_no=clean_text(first.get('Account Number'))
    account_last4=clean_text(first.get('Account Last 4'))
    if not account_last4 and account_no:
        account_last4=re.sub(r'\D','',account_no)[-4:]

    file_stub=f"{seq:02d}_{safe_file_component(bank_name,35)}"
    if account_last4:
        file_stub += f"_{safe_file_component(account_last4,8)}"
    out_path=os.path.join(BANK_OUTPUT_DIR,file_stub+'.xlsx')

    wb=Workbook(); wb.remove(wb.active)

    # COVER
    cover_rows=pd.DataFrame([
        ['Bank',bank_name],
        ['Account Number',account_no],
        ['Source File',source_file],
        ['Transactions',len(grp)],
        ['Identified Customers',int(grp['Customer Name'].fillna('').astype(str).str.strip().replace('',np.nan).nunique())],
        ['Unidentified Customer Rows',int(grp['Customer Name'].fillna('').astype(str).str.strip().eq('').sum())],
        ['Bill Payment Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('BILL PAYMENT').sum())],
        ['NEFT Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('NEFT').sum())],
        ['IMPS Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('IMPS').sum())],
        ['UPI Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('UPI').sum())],
        ['RTGS Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('RTGS').sum())],
        ['Cash Deposit Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('CASH DEPOSIT').sum())],
        ['Cash Withdrawal Transactions',int(add_v851_audit_columns(grp)['Payment Mode'].eq('CASH WITHDRAWAL').sum())],
        ['Total Actual Charges / Fees',float(add_v851_audit_columns(grp)['Charge Amount'].fillna(0).sum())],
        ['Total Debit',float(grp['Debit'].fillna(0).sum())],
        ['Total Credit',float(grp['Credit'].fillna(0).sum())],
        ['Net Credit - Debit',float(grp['Credit'].fillna(0).sum()-grp['Debit'].fillna(0).sum())],
        ['First Date',grp['Date'].dropna().min() if grp['Date'].notna().any() else pd.NaT],
        ['Last Date',grp['Date'].dropna().max() if grp['Date'].notna().any() else pd.NaT],
        ['Author',AUTHOR_NAME],
        ['Instagram',AUTHOR_INSTAGRAM],
    ],columns=['Metric','Value'])
    add_df_sheet(wb,'COVER',cover_rows)

    # MONTH-WISE — exactly the requested view for this bank/source only.
    monthly=month_summary_for_group(grp)
    add_df_sheet(wb,'MONTHLY_SUMMARY',monthly,
                 money_cols=('Total Debit','Total Credit','Net Credit - Debit','Month End Available Balance'))

    # V8.5 — CUSTOMER-FIRST REPORTING
    customer_summary = customer_wise_summary_for_group(grp)
    add_df_sheet(
        wb,'CUSTOMER_WISE_SUMMARY',customer_summary,
        money_cols=('Total Debit','Total Credit','Net Credit - Debit'),
        date_cols=('First Transaction Date','Last Transaction Date')
    )

    customer_monthly = customer_monthly_summary_for_group(grp)
    add_df_sheet(
        wb,'CUSTOMER_MONTHLY',customer_monthly,
        money_cols=('Total Debit','Total Credit','Net Credit - Debit'),
        date_cols=('First Transaction Date','Last Transaction Date')
    )

    customer_ledger = customer_ledger_for_group(grp)
    add_df_sheet(
        wb,'CUSTOMER_LEDGER',customer_ledger,
        money_cols=('Debit','Credit','Balance','Transaction Amount'),
        date_cols=('Date',)
    )

    customer_control = customer_control_for_group(grp, customer_summary, customer_monthly)
    add_df_sheet(
        wb,'CUSTOMER_CONTROL',customer_control,
        money_cols=()
    )

    # V8.5.1 — EASY AUDIT MODE + CHARGE REPORTS
    audit_grp = add_v851_audit_columns(grp)

    mode_summary = mode_summary_v851(grp)
    add_df_sheet(
        wb,'MODE_SUMMARY',mode_summary,
        money_cols=('Total Debit','Total Credit','Net Credit - Debit'),
        highlight_modes=True
    )

    charges_summary = charges_summary_v851(grp)
    add_df_sheet(
        wb,'CHARGES_SUMMARY',charges_summary,
        money_cols=('Net Charge Amount','Total Debit Charges','Total Credit/Reversal'),
        date_cols=('First Date','Last Date'),
        highlight_charges=True
    )

    charge_ledger = charge_ledger_v851(grp)
    add_df_sheet(
        wb,'CHARGES_LEDGER',charge_ledger,
        money_cols=('Charge Amount','Bank Charges / Fees','GST / Tax Charges','Bill Payment Fees',
                    'Debit','Credit','Balance'),
        date_cols=('Date',),
        highlight_charges=True
    )

    # EASY STATEMENT — clean normalized statement for this source only.
    grp = audit_grp
    easy_cols=[c for c in [
        'Date','Customer Name','UTR / Reference','Narration',
        'Payment Mode','Charge Type','Is Charge / Fee',
        'Bank Charges / Fees','GST / Tax Charges','Bill Payment Fees','Charge Amount',
        'Debit','Credit','Balance','Direction','Transaction Amount',
        'Transaction Type','Possible Duplicate','Review Status',
        'Source Part','Source Row'
    ] if c in grp.columns]
    add_df_sheet(
        wb,'EASY_STATEMENT',grp[easy_cols],
        money_cols=('Bank Charges / Fees','GST / Tax Charges','Bill Payment Fees','Charge Amount',
                    'Debit','Credit','Balance','Transaction Amount'),
        date_cols=('Date',),
        highlight_modes=True,
        highlight_charges=True
    )

    # AUDIT CONTROL for this source only.
    recon=bank_reconciliation[bank_reconciliation['Source File'].astype(str).eq(str(source_file))].copy() if 'Source File' in bank_reconciliation.columns else pd.DataFrame()
    add_df_sheet(wb,'RECONCILIATION',recon,
                 money_cols=('Derived Opening Balance','Total Debit','Total Credit','Expected Closing Balance','Actual Closing Balance','Difference'),
                 date_cols=('First Date','Last Date'))

    # INTERBANK candidate matches involving this source only.
    if interbank_matches.empty:
        source_inter=interbank_matches
    else:
        source_inter=interbank_matches[
            interbank_matches['Debit Source File'].astype(str).eq(str(source_file)) |
            interbank_matches['Credit Source File'].astype(str).eq(str(source_file))
        ].copy()
    add_df_sheet(wb,'INTERBANK_MATCHES',source_inter,
                 money_cols=('Amount',),date_cols=('Debit Date','Credit Date'),highlight_review=True)

    # REAL RAW DATA — every raw sheet/table belonging to this exact source file.
    raws=raw_by_source.get(source_file,[])
    if raws:
        for raw_no,(part,raw_name,raw_df) in enumerate(raws,1):
            part_label=safe_file_component(part,16) or f'PART_{raw_no:02d}'
            add_raw_sheet(wb,f'REAL_RAW_{raw_no:02d}_{part_label}',raw_df)
    else:
        ws=wb.create_sheet('REAL_RAW_NOT_CAPTURED')
        ws['A1']='Raw data was not captured for this run.'
        ws['A2']='Ensure config.yaml contains: preserve_raw_data: true'

    wb.properties.creator=AUTHOR_NAME
    wb.properties.lastModifiedBy=AUTHOR_NAME
    wb.properties.title=f'{bank_name} Bank Statement Report {DISPLAY_VERSION}'
    wb.properties.subject='Bank-wise customer summary, customer monthly debit-credit, customer ledger, raw statement and audit controls'
    wb.properties.description=f'Prepared by {AUTHOR_NAME} | {AUTHOR_INSTAGRAM} | Source: {source_file}'
    wb.save(out_path)
    generated_files.append(out_path)
    logger.info('BANK FILE CREATED: %s',out_path)

print('')
print('='*82)
print('SUCCESS — BANK-WISE SEPARATE REPORTS V8.5.1 AUDIT MODES + CHARGES')
print('='*82)
print('Output Folder :',BANK_OUTPUT_DIR)
print('Bank Files    :',len(generated_files))
print('Transactions  :',f'{len(transactions):,}')
print('Interbank Candidates:',f'{len(interbank_matches):,}')
print('')
for p in generated_files:
    print(' -',os.path.basename(p))
print('')
print('Every bank/source workbook contains:')
print('  COVER | MONTHLY_SUMMARY | CUSTOMER_WISE_SUMMARY | CUSTOMER_MONTHLY | CUSTOMER_LEDGER | CUSTOMER_CONTROL | MODE_SUMMARY | CHARGES_SUMMARY | CHARGES_LEDGER | EASY_STATEMENT | RECONCILIATION | INTERBANK_MATCHES | REAL_RAW_*')
print('')
print('INTERBANK_MATCHES are CANDIDATES ONLY and must be verified from the actual bank statements.')
print('Author:',AUTHOR_NAME,'|',AUTHOR_INSTAGRAM)
print('='*82)
